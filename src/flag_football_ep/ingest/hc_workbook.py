"""Head-coach workbook ingest: cell-level `.xlsx` reading, dtype-based block
segmentation, and per-block contract mapping for the three hand-maintained
Excel workbooks under `data/raw/hc_files/` (gitignored, PII -- player names).

Why cell-level reading instead of `pl.read_excel`: one sheet is not one
table. `Scoring Probability by Situation 2023-2026.xlsx`'s `Data` and
`Copy of Data` tabs carry two incompatible row layouts under a single header
row -- roughly one row in six holds a team-name pair where the header claims
`PLAY #`/`ODK`, the rest hold a real numeric `PLAY #`. `Germany Analytics
Stats EC 2025 vs WC Nations.xlsx`'s `Data` tab carries a correct header over
thousands of completely empty rows. A bulk rectangular reader (calamine/
`fastexcel`) assumes one consistent layout per sheet and cannot represent
either finding; both must come out of this module as named, counted
findings instead of a silently corrupted frame.

Order of operations this module implements: `read_sheet_rows` (openpyxl,
`data_only=True` formula resolution + `read_only=True` streaming) ->
`segment_blocks` (classify every row by its first cell's dtype, before any
column mapping is trusted -- M3-01-RESEARCH.md Pattern 1) ->
`map_block_to_frame` (per block: normalize the header, cast every cell to
Utf8, null out the pair block's unresolved tail, materialize absent
contract core columns, `validate_header`/`check_column_domains`, rename
charting columns onto their canonical extras -- Pattern 2). Nothing in this
module raises on a data-quality finding; `SheetNotFoundError` is the only
exception it defines, for a structurally absent sheet. Everything else is
folded into the caller's `HcIngestNotices.messages`.

Game segmentation and identity: `segment_games` splits a block into
per-game slices (numeric: `PLAY #` reset; pair: team-pair change);
`resolve_game_identity` resolves a slice against the maintained
`data/reference/hc_games.csv`, degrading to a provisional id plus a notice
on a miss rather than raising (HC-D04).

`ingest_workbook` is this module's convergence point: it reuses
`ingest.hudl`'s derivation chain unchanged (`derive_identity_columns`,
`parse_result_tokens`, `derive_outcome_columns`, `derive_drive_id`,
`derive_yards_gained_first_down`) rather than forking a second
RESULT-parsing implementation (HC-D01), and calls
`canonical.conform_to_canonical` to converge onto the canonical schema with
`source = hc_workbook:<file>:<sheet>`.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
import polars as pl

from flag_football_ep.canonical import (
    CORE_COLUMNS,
    NULLABLE_EXTRAS,
    ConformReport,
    add_score_columns,
    add_scoring_play_team,
    conform_to_canonical,
)
from flag_football_ep.ingest import hudl
from flag_football_ep.reference import map_players
from flag_football_ep.validation.schema import (
    Contract,
    DomainViolation,
    HeaderReport,
    check_column_domains,
    validate_header,
)

__all__ = [
    "SHEET_NAMES",
    "PAIR_BLOCK_TAIL_ANCHOR",
    "HALF_SENTINEL",
    "HALF_SENTINEL_EXCLUDED_SHEETS",
    "SheetNotFoundError",
    "slugify",
    "hc_source_label",
    "HcBlock",
    "HcIngestNotices",
    "read_sheet_rows",
    "segment_blocks",
    "map_block_to_frame",
    "HcGameSlice",
    "HcGameIdentity",
    "segment_games",
    "resolve_game_identity",
    "count_result_tokens",
    "ingest_workbook",
]

# The two tab names every HC workbook charts plays under. `Copy of Data`
# (Scoring Probability workbook) is very likely an Excel-generated snapshot
# of `Data`, not an independent charting session (M3-01-RESEARCH.md
# Pitfall 4) -- dedupe between the two is a later plan's job, not this one's.
SHEET_NAMES = ("Data", "Copy of Data")

# By header position (not by name -- the header names past this point are
# exactly what is unknown for a pair block), the column at which a pair
# block's semantics become unresolved. See map_block_to_frame.
PAIR_BLOCK_TAIL_ANCHOR = "RECEIVED BY"

# The constant `half` sentinel stamped on every row of a declared,
# non-`Copy of Data` HC game (see ingest_workbook). `2` -- not `null` and not
# `1` -- is the only value that both (a) genuinely satisfies
# `validation.checks.half_assigned`'s contract (`half in {1, 2}`), and (b)
# keeps `features.mutations._mark_half_end`'s `game_end` (which requires
# `half == 2`) firing exactly once, at the true last row -- which is what
# resolves WP's `Winner` and EP's post-game `None`-ing. M3-02-RESEARCH.md
# Sec 2.2 tabulates why `null`/`1`/any other value all fail one or both of
# these. This value must never be inlined anywhere else in this module.
HALF_SENTINEL = 2

# `HALF_SENTINEL` is never stamped for rows from these sheets, even for a
# game declared in `data/reference/hc_games.csv` -- `Copy of Data`'s column
# layout differs from `Data`'s in ways not yet resolved (Frage 2,
# M3-02-RESEARCH.md Sec 1.3: 14 vs 15 columns, `YARD LN`/`Drive Success`
# swapped, extra `FH`, `Thrown By`/`YAC` absent). Its rows keep `half = null`
# and keep quarantining on `half_assigned` until Frage 2 is answered.
HALF_SENTINEL_EXCLUDED_SHEETS = ("Copy of Data",)

# Extends hudl._CHARTING_RENAME with the HC-only charting columns that have
# no equivalent in any Hudl export (HC-D01: reuse, don't fork). Several of
# these (OFF STR, THROWN BY, YAC) already exist in hudl._CHARTING_RENAME --
# re-declared here for documentation purposes; the dict-merge below is a
# harmless no-op overwrite for those keys. Matched case-insensitively and
# whitespace-trimmed against the sheet header (see _rename_target).
_HC_ONLY_RENAME: dict[str, str] = {
    "AIR YARDS": "air_yards",
    "BF ACTION": "bf_action",
    "HAND": "hand",
    "EFFICIENCY": "efficiency",
    "DRIVE SUCCESS": "drive_success",
    "OFF STR": "off_str",
    "THROWN BY": "thrown_by",
    "YAC": "yac",
}
_HC_RENAME: dict[str, str] = {**hudl._CHARTING_RENAME, **_HC_ONLY_RENAME}
# case-insensitive/whitespace-tolerant lookup: normalized header -> canonical extra
_HC_RENAME_UPPER: dict[str, str] = {k.strip().upper(): v for k, v in _HC_RENAME.items()}

# Columns this module invents for the pair block's team-name-pair cells
# (plan M3-01-03 uses them as the game-identity key) -- never flagged as an
# unmapped header, since they are not sheet headers at all.
_SYNTHETIC_COLUMNS = frozenset({"hc_pair_team1", "hc_pair_team2"})

# The three ODK-marker letters the head coach used in a pair block's column A
# for a while, instead of writing a team-name pair on every row -- confirmed
# 2026-09-03 (docs/hc-rueckfragen-2026-09.md Frage 2, Antwort): "O" =
# Offense, "D" = Defense, "S" = kein echter Play ("no-play", handled like
# Timeout/Offsetting Penalties -- see the ODK == "S" override in
# ingest_workbook). Matched case-insensitively, whitespace-trimmed, against
# column A (index 0) only: the head coach's own wording ("column 1") and the
# confirmed header-row layout (team1/team2 also sit in columns A/B) both
# point at column A as the marker's position, not column B.
_PAIR_MARKER_VALUES = frozenset({"O", "D", "S"})


class SheetNotFoundError(Exception):
    """Raised when the requested sheet is absent from the workbook."""


@dataclass(frozen=True)
class HcBlock:
    """One contiguous run of same-kind rows within a sheet (block
    segmentation by column-1 dtype, not by header text -- see module
    docstring Pattern 1).

    `index` is the 0-based order of appearance within the sheet. `header` is
    the sheet's single header row, shared by every block (both block kinds
    are read under the same physical header row -- that mismatch for `pair`
    blocks is exactly the finding this module exists to report, not hide).
    `rows` is `(physical_row_number, cell_values)` per row, in sheet order.
    """

    index: int
    kind: str  # "pair" | "numeric"
    header: list[Any]
    rows: list[tuple[int, tuple]]
    first_row: int
    last_row: int


@dataclass
class HcIngestNotices:
    """Everything found for one (workbook, sheet): never raised, always
    returned. Keyed on the source label rather than a game id (mirrors
    `hudl.IngestNotices`), because one HC sheet holds many games across
    possibly several blocks -- there is no single game id to key on until
    game-identity resolution (`resolve_game_identity`) runs.

    `unmapped_players` is PII (raw player names/jersey labels left unmapped
    by `reference.map_players`) -- it exists so a caller can decide what to
    do with the labels themselves (e.g. extend `player_mapping.csv`), but it
    must NEVER be rendered into a report, a console line, a doc or a commit
    message; only its length (`len(notices.unmapped_players)`) may ever
    appear in human-readable output. `result_token_counts` is not PII: every
    token observed in the sheet's `RESULT` column (contract vocabulary or
    not), counted by `count_result_tokens`.
    """

    source_label: str
    sheet: str
    messages: list[str] = field(default_factory=list)
    header: HeaderReport | None = None
    domain: list[DomainViolation] = field(default_factory=list)
    conform: ConformReport | None = None
    unmapped_players: list[str] = field(default_factory=list)
    result_token_counts: dict[str, int] = field(default_factory=dict)


def slugify(value: str) -> str:
    """Lowercase; every run of non-alphanumeric characters becomes a single
    `-`; strip leading/trailing `-`."""
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", value.lower())
    return slug.strip("-")


def hc_source_label(path: Path, sheet: str) -> str:
    """The value that lands in the canonical `source` column (HC-D01):
    `hc_workbook:<slugified file stem>:<slugified sheet name>`.
    """
    return f"hc_workbook:{slugify(Path(path).stem)}:{slugify(sheet)}"


def read_sheet_rows(path: Path, sheet: str) -> tuple[list[Any], list[tuple[int, tuple]], list[str]]:
    """Read one sheet's header and data rows with formula resolution.

    `openpyxl.load_workbook(path, data_only=True, read_only=True)`; always
    closes the workbook in a `finally` (read-only workbooks otherwise keep a
    file handle open). Header comes from row 1; data rows from
    `ws.iter_rows(min_row=2, values_only=True)`, physical row number kept
    alongside each kept row.

    Rows where every cell is `None` or an empty string are skipped and
    counted (a blank row inside the sheet's used range is not data). Cells
    holding the literal string `"#N/A"` (unresolved formula residue) are
    counted. A sheet whose header is intact but whose data rows are all
    blank is reported via a message naming both the physical row count and
    the (zero) kept count -- callers must be able to tell "empty tab" from
    "no such sheet" (`SheetNotFoundError`, raised only when `sheet` is
    absent from `wb.sheetnames`).
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise SheetNotFoundError(
                f"sheet {sheet!r} not found in {Path(path).name!r}; "
                f"available sheets: {wb.sheetnames}"
            )
        ws = wb[sheet]

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        header = list(header_row)

        rows: list[tuple[int, tuple]] = []
        physical = 0
        blank = 0
        na_count = 0
        for row_num, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            physical += 1
            if all(v is None or v == "" for v in values):
                blank += 1
                continue
            na_count += sum(1 for v in values if v == "#N/A")
            rows.append((row_num, values))

        messages: list[str] = []
        kept = len(rows)
        if kept == 0:
            messages.append(
                f"Sheet {sheet!r} ist leer/empty: {physical} physische Zeile(n) gesehen, "
                f"{kept} Datenzeile(n) übernommen -- als leere Quelle melden, nicht als "
                "Quelle mit null Spielen behandeln"
            )
        if blank:
            messages.append(f"{blank} leere Zeile(n) innerhalb des Bereichs übersprungen")
        if na_count:
            messages.append(f"{na_count} Zelle(n) mit '#N/A'-Formelresten gefunden")

        return header, rows, messages
    finally:
        wb.close()


def segment_blocks(
    header: list[Any], rows: list[tuple[int, tuple]]
) -> tuple[list[HcBlock], list[str]]:
    """Classify every row by its first cell and group consecutive same-kind
    rows into blocks, in sheet order (Pattern 1: block segmentation by
    column-1 dtype, not by header text).

    `isinstance(v, (int, float)) and not isinstance(v, bool)` is `numeric`
    (openpyxl always delivers Excel numbers as `1.0`, never `1` -- Pitfall
    5; `bool` is an `int` subclass in Python, excluded explicitly so a
    charted `True`/`False` cell never misclassifies as numeric). A non-empty
    `str` first cell is `pair`. Anything else (`None`, `""`, or any other
    type) is a skipped row -- counted, but never treated as a block
    boundary, so a stray blank first cell in the middle of a run does not
    fracture one real block into two.
    """
    blocks: list[HcBlock] = []
    current_kind: str | None = None
    current_rows: list[tuple[int, tuple]] = []
    skipped = 0

    def _flush() -> None:
        if current_kind is None or not current_rows:
            return
        blocks.append(
            HcBlock(
                index=len(blocks),
                kind=current_kind,
                header=header,
                rows=list(current_rows),
                first_row=current_rows[0][0],
                last_row=current_rows[-1][0],
            )
        )

    for row_num, values in rows:
        first_cell = values[0] if values else None
        if isinstance(first_cell, bool):
            kind: str | None = None
        elif isinstance(first_cell, (int, float)):
            kind = "numeric"
        elif isinstance(first_cell, str) and first_cell != "":
            kind = "pair"
        else:
            kind = None

        if kind is None:
            skipped += 1
            continue

        if kind != current_kind:
            _flush()
            current_kind = kind
            current_rows = []
        current_rows.append((row_num, values))

    _flush()

    messages: list[str] = []
    if skipped:
        messages.append(
            f"{skipped} Zeile(n) bei der Blockerkennung übersprungen "
            "(erste Zelle weder Zahl noch Text)"
        )
    for block in blocks:
        messages.append(
            f"Block {block.index} ({block.kind}): rows {block.first_row}-{block.last_row}, "
            f"{len(block.rows)} Zeilen"
        )
    if blocks:
        n_pair = sum(len(b.rows) for b in blocks if b.kind == "pair")
        n_numeric = sum(len(b.rows) for b in blocks if b.kind == "numeric")
        total = n_pair + n_numeric
        messages.append(f"Block-Split: {n_pair}/{total} pair, {n_numeric}/{total} numeric")

    return blocks, messages


def _normalize_header(header: list[Any]) -> tuple[list[str], list[int], list[str]]:
    """Trim header names, drop `None`/empty ones, de-duplicate the rest.

    Returns `(clean_names, kept_indices, messages)`: `clean_names[i]` is the
    (possibly de-duplicated) name for the original column at `kept_indices[i]`.
    A repeated name gets `_2`, `_3`, ... appended for its 2nd, 3rd, ...
    occurrence; each rename is named in a message so a duplicate charting
    column is never silently merged into the first one under the same name.
    """
    messages: list[str] = []
    seen: dict[str, int] = {}
    clean_names: list[str] = []
    kept_indices: list[int] = []

    for i, raw_name in enumerate(header):
        if raw_name is None:
            continue
        name = str(raw_name).strip()
        if name == "":
            continue

        seen[name] = seen.get(name, 0) + 1
        if seen[name] > 1:
            deduped = f"{name}_{seen[name]}"
            messages.append(
                f"doppelte Spalte {name!r} an Position {i} als {deduped!r} umbenannt"
            )
            name = deduped

        clean_names.append(name)
        kept_indices.append(i)

    return clean_names, kept_indices, messages


def _cell_to_utf8(value: Any, *, strip_integral: bool) -> str | None:
    """Cast one cell to its Utf8 representation.

    `None` stays null. A `bool` (checked before the numeric branch -- `bool`
    is an `int` subclass) becomes its Python str form. When `strip_integral`
    is true, a `float` with no fractional part becomes its plain integer
    string (`25.0` -> `"25"`) -- required both so jersey-number-shaped
    columns are usable as `player_mapping.csv` lookup keys, and so a
    genuinely numeric contract column (DN/DIST/YARD LN/PLAY #) stays
    castable downstream (a trailing ".0" would fail polars' non-strict
    str->int cast). `RESULT` is the one column this module calls with
    `strip_integral=False`: it is a free-text contract column, so a numeric
    charting error landing there (the real `-5.0` found in the corpus) is
    preserved verbatim as evidence of the error rather than normalized away.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and strip_integral and value.is_integer():
        return str(int(value))
    return str(value)


def _rename_target(header_name: str) -> str | None:
    """The canonical extras name for `header_name`, or `None` if this module
    has no rename for it (case-insensitive, whitespace-trimmed lookup)."""
    return _HC_RENAME_UPPER.get(header_name.strip().upper())


def _null_pair_block_tail(
    df: pl.DataFrame, clean_names: list[str], marker_odk: list[str | None]
) -> tuple[pl.DataFrame, list[str]]:
    """Pair-block handling (M3-01-RESEARCH.md Pitfall 2 / Open Question #2).

    The first two columns hold either a team-name pair or (Frage 2, Antwort
    2026-09-03) an O/D/S marker in column A -- their raw values survive
    under dedicated names (`hc_pair_team1`/`hc_pair_team2`, plan M3-01-03's
    game-identity key) before `PLAY #`/`ODK` themselves are overwritten.
    `PLAY #` is nulled for every row (no real play numbering exists in
    either style; reassigned later by `_reassign_hc_play_no`, HC corpus
    admission rule 2, 2026-09-04). `ODK`
    is nulled for a team-name header row (still no real ODK there) but set
    to `marker_odk[i]` for a marker row -- `_pair_row_marker` already
    confirmed that value is one of `"O"`/`"D"`/`"S"`, so this is not a
    guess, it is the row's own charted marker. Columns from
    `PAIR_BLOCK_TAIL_ANCHOR` onward (by position, not by name -- the header
    names past that point are exactly what is unknown for this block) are
    nulled with one notice naming the reason: guessing the column shift
    would swap passer, receiver and gain, and Frage 2's answer addressed the
    block-segmentation/ODK question, not this tail-column question, so the
    same conservative nulling still applies to marker rows too. Columns
    before the anchor (through `TARGET ROUTE`) are left untouched -- they
    line up with the header in both a team-name and a marker row
    (M3-01-RESEARCH.md Pitfall 2).
    """
    messages: list[str] = []
    n_rows = df.height

    if len(clean_names) >= 1:
        df = df.with_columns(pl.col(clean_names[0]).alias("hc_pair_team1"))
    if len(clean_names) >= 2:
        df = df.with_columns(pl.col(clean_names[1]).alias("hc_pair_team2"))

    if "PLAY #" in df.columns:
        df = df.with_columns(pl.lit(None, dtype=pl.Utf8).alias("PLAY #"))

    n_marker = sum(1 for v in marker_odk if v is not None)
    if "ODK" in df.columns:
        if n_rows and len(marker_odk) == n_rows:
            df = df.with_columns(pl.Series("ODK", marker_odk, dtype=pl.Utf8))
        else:
            df = df.with_columns(pl.lit(None, dtype=pl.Utf8).alias("ODK"))
    messages.append(
        f"Pair-Block: PLAY # für {n_rows} Zeile(n) auf null gesetzt (Spalte "
        "enthält stattdessen ein Team-Namenspaar oder eine Marker-Zelle); ODK "
        f"für {n_marker} Marker-Zeile(n) aus O/D/S übernommen (Frage 2, Antwort "
        f"2026-09-03), für die übrigen {n_rows - n_marker} Zeile(n) (Team-"
        "Namenspaar-Kopfzeile) auf null gesetzt"
    )

    anchor_idx = next(
        (
            i
            for i, name in enumerate(clean_names)
            if name.strip().upper() == PAIR_BLOCK_TAIL_ANCHOR.upper()
        ),
        None,
    )
    if anchor_idx is not None:
        tail_names = clean_names[anchor_idx:]
        if tail_names:
            df = df.with_columns(
                [pl.lit(None, dtype=pl.Utf8).alias(name) for name in tail_names]
            )
            messages.append(
                f"Pair-Block: Spalten ab {PAIR_BLOCK_TAIL_ANCHOR!r} ({len(tail_names)} "
                f"Spalte(n), {n_rows} Zeile(n)) auf null gesetzt -- Frage 2 offen "
                "(Spaltenversatz ungeklärt); ein geratener Spaltenversatz würde "
                "Passgeber, Empfänger und Raumgewinn vertauschen"
            )

    return df, messages


def map_block_to_frame(
    block: HcBlock, contract: Contract
) -> tuple[pl.DataFrame, HeaderReport, list[DomainViolation], list[str]]:
    """Map one `HcBlock` onto the contract's raw column names, dtype-validated.

    Order of operations (Pattern 2: dtype-validated mapping, never
    header-text-only):
    1. Normalize the header (`_normalize_header`).
    2. Build the frame with every cell cast to Utf8 (`_cell_to_utf8`).
    3. For a `pair` block, null out the unresolved tail (`_null_pair_block_tail`).
    4. Materialize every absent contract core column as all-null Utf8 (every
       HC sheet lacks `PLAY TYPE`; a pair block also lacks a real `PLAY #`/
       `ODK`, but those are already present as columns by this point --
       nulled, not absent). Then `validate_header` (which now cannot raise,
       since every core column already exists) and `check_column_domains`.
    5. Rename charting columns onto their canonical extras via `_HC_RENAME`;
       every header with no contract slot and no rename target is collected
       into one notice, never silently dropped.

    Never raises on a data-quality finding; a block whose header carries no
    usable column names returns an empty frame plus a message.
    """
    messages: list[str] = []

    clean_names, kept_indices, dedup_messages = _normalize_header(block.header)
    messages.extend(dedup_messages)

    if not clean_names:
        messages.append(
            f"Block {block.index} ({block.kind}): keine verwertbaren Spalten im Header, "
            "leerer Frame zurückgegeben"
        )
        return (
            pl.DataFrame(),
            HeaderReport(missing_core=[], materialized_optional=[], unknown=[]),
            [],
            messages,
        )

    columns_data: dict[str, list[str | None]] = {name: [] for name in clean_names}
    for _row_num, values in block.rows:
        for name, idx in zip(clean_names, kept_indices):
            value = values[idx] if idx < len(values) else None
            strip_integral = name != "RESULT"
            columns_data[name].append(_cell_to_utf8(value, strip_integral=strip_integral))

    df = pl.DataFrame(columns_data, schema={name: pl.Utf8 for name in clean_names})

    if block.kind == "pair":
        marker_odk = [
            _pair_row_marker(
                values[0] if len(values) >= 1 else None,
                values[1] if len(values) >= 2 else None,
            )
            for _, values in block.rows
        ]
        df, pair_messages = _null_pair_block_tail(df, clean_names, marker_odk)
        messages.extend(pair_messages)

    materialized_core = [c for c in contract.core_columns if c not in df.columns]
    if materialized_core:
        df = df.with_columns(
            [pl.lit(None, dtype=pl.Utf8).alias(c) for c in materialized_core]
        )
        messages.append(
            f"Kernspalte(n) im Sheet nicht vorhanden, als null angelegt: {materialized_core}"
        )

    df, header_report = validate_header(df, contract)
    domain_violations = check_column_domains(df, contract)

    known_contract = set(contract.core_columns) | set(contract.optional_columns)
    rename_map: dict[str, str] = {}
    unmapped: list[str] = []
    for name in df.columns:
        if name in _SYNTHETIC_COLUMNS or name in known_contract:
            continue
        target = _rename_target(name)
        if target is not None:
            rename_map[name] = target
        else:
            unmapped.append(name)

    if unmapped:
        messages.append(
            f"Ohne kanonisches Ziel, nicht stillschweigend verworfen: {unmapped}"
        )
    if rename_map:
        df = df.rename(rename_map)

    return df, header_report, domain_violations, messages


@dataclass(frozen=True)
class HcGameSlice:
    """One game's worth of rows inside a single `HcBlock` (game segmentation
    within a block -- see `segment_games`).

    `block_index`/`game_index` are 0-based; `block_key` is the stable,
    block-scoped identifier `b{block_index:02d}-g{game_index:02d}` that
    `data/reference/hc_games.csv` keys on. `rows` is the game's row slice in
    the same `(physical_row_number, cell_values)` shape as `HcBlock.rows`.
    `source_team1`/`source_team2` are the raw team-pair labels for a `pair`
    slice (from the slice's first row), `None` for a `numeric` slice.
    """

    block_index: int
    game_index: int
    block_key: str
    kind: str  # "pair" | "numeric"
    rows: list[tuple[int, tuple]]
    first_row: int
    last_row: int
    source_team1: str | None
    source_team2: str | None


def _coerce_play_number(value: Any) -> float | None:
    """Best-effort numeric coercion of a `PLAY #` cell (Pitfall 5: openpyxl
    delivers Excel numbers as `float`). `None`/non-numeric/unparseable
    strings become `None` -- treated by `segment_games` as "does not
    increase", i.e. a game boundary, never silently merged into the
    previous game.
    """
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.strip())
        except ValueError:
            return None
    return None


def _normalize_pair_label(value: Any) -> str:
    """Case-insensitive, whitespace-trimmed comparison key for a team-pair
    label. The raw value (not this normalized form) is what survives into
    `HcGameSlice.source_team1`/`source_team2`.
    """
    if value is None:
        return ""
    return str(value).strip().casefold()


def _pair_row_marker(value: Any, second_value: Any = None) -> str | None:
    """`"O"`/`"D"`/`"S"` if `value` (a pair block's column-A cell) is one of
    the head coach's ODK markers AND `second_value` (column B) is empty,
    else `None`.

    The `second_value` check exists because `"D"`/`"S"` are ALSO genuine,
    real abbreviated team-name values in this corpus (`"D"` = Deutschland,
    confirmed 6 times in the real `Data`-tab pair block; `"S"`/`"K"` occur
    too) -- `_pair_row_marker` cannot tell "D" (Defense marker) from "D"
    (Deutschland) from the letter alone, per RESEARCH.md Sec 1.2's own
    warning against guessing ambiguous abbreviations. Every real occurrence
    of `"D"`/`"S"`/`"K"` found in column A of the real workbook (2026-09-03)
    carries a second, non-empty team-abbreviation value in column B (e.g.
    `("D", "AT")`, `("S", "S")`, `("K", "K")`) -- a GENUINE marker row, per
    the head coach's own description, carries no second team name (nothing
    else to write). Requiring `second_value` to be empty is therefore not a
    guess, it is the only reading that matches every real example of both
    cases: an earlier version of this function ignored column B entirely and
    mis-classified those 6 real `"D"` abbreviation rows as markers, silently
    changing the real `Data`-tab pair block's game count from the validated
    137 -> 22 (M3-02-RESEARCH.md Sec 1.2) to 137 -> 16.
    """
    if value is None:
        return None
    if second_value is not None and str(second_value).strip() != "":
        return None
    normalized = str(value).strip().upper()
    return normalized if normalized in _PAIR_MARKER_VALUES else None


def _split_numeric_block(
    rows: list[tuple[int, tuple]], messages: list[str]
) -> list[list[tuple[int, tuple]]]:
    """Split a numeric block's rows into games wherever `PLAY #` (column 0)
    does not increase relative to the previous row -- a reset to 1, any
    decrease, or a null/unparseable cell, which is itself always a boundary
    (never compared against a stale reference value, so one corrupt cell
    cannot make two real games silently look like one)."""
    groups: list[list[tuple[int, tuple]]] = []
    current: list[tuple[int, tuple]] = []
    prev_play_num: float | None = None
    unparseable = 0

    for row_num, values in rows:
        raw = values[0] if values else None
        play_num = _coerce_play_number(raw)
        if play_num is None:
            unparseable += 1

        is_boundary = (
            not current or play_num is None or prev_play_num is None or play_num <= prev_play_num
        )
        if is_boundary and current:
            groups.append(current)
            current = []
        current.append((row_num, values))
        prev_play_num = play_num

    if current:
        groups.append(current)

    if unparseable:
        messages.append(
            f"{unparseable} Zeile(n) mit nicht auswertbarer PLAY # (null oder nicht "
            "parsebar) -- als Spielgrenze behandelt statt zwei Spiele stillschweigend "
            "zusammenzuführen"
        )

    return groups


def _split_pair_block(
    rows: list[tuple[int, tuple]],
) -> list[tuple[list[tuple[int, tuple]], str | None, str | None]]:
    """Split a pair block's rows into games, honouring the two charting
    conventions confirmed for `hc_workbook`'s pair blocks (Frage 2,
    `docs/hc-rueckfragen-2026-09.md`, Antwort 2026-09-03):

    1. **Team-name-per-row era** (M3-02-RESEARCH.md Sec 1.2 -- the only style
       found so far in the real `Data` tab's pair block): the head coach
       charts offense/defense possessions of the SAME game as flipped
       team-pair rows (`Germany | Ireland` -> `Ireland | Germany` -> ...).
       Consecutive rows whose unordered `{team1, team2}` pair (case-
       insensitive, whitespace-trimmed) stays the same, with no O/D/S marker
       row in between, stay one game -- comparing the
       *ordered* pair would treat every possession flip as a new game,
       fragmenting one real game into dozens of slices (137 -> 22 on the real
       block, M3-02-RESEARCH.md Sec 1.2).
    2. **Header + O/D/S-marker era** (the head coach's Antwort: "Ich habe
       irgendwann aufgehört die Teamnamen aufzuschreiben und nur noch O für
       Offense, D für Defense und S für no-play genommen. Dann wiederum
       irgendwann wieder angefangen zumindest in der ersten Zeile wieder die
       Teamnamen niederzuschreiben. Alles was darunter kommt, bis zu einer
       leeren Zeile bzw. einer neuen Zeile mit Teamnamen soll diese Teams
       darstellen."): a team-name row opens a block; every `_pair_row_marker`
       row after it (in column A) belongs to that block, carrying no team
       name of its own; the block ends at the next team-name row. Once a
       block has received a marker row it is "closed" to further same-pair
       merging under rule 1 above -- a repeated team-name row after markers
       always opens a NEW block (rule 2's literal boundary), never merges
       back into the marker block.

       **Known, deliberate gap:** the head coach's rule also names a blank
       row as a boundary. This function does NOT detect that: by the time a
       block's rows reach here, `read_sheet_rows` has already stripped every
       genuinely blank row (no trace survives to compare against), and
       `segment_blocks` separately, silently skips any row whose column A is
       neither numeric nor a non-empty string -- a real, populated play row
       with an empty column A (5 such rows found in the real `Data`-tab pair
       block, 2026-09-03: e.g. row 43, DN/DIST/YARD LN populated, column A/B
       both empty) is indistinguishable, from inside this function, from a
       genuinely blank row. An earlier version of this function inferred a
       blank-row boundary from a gap in physical row numbers; empirically
       verified against the real workbook, that inference is WRONG -- it
       over-fragmented the real block from the validated 22 games (M3-02-
       RESEARCH.md Sec 1.2) to 18, entirely from those 5 non-blank
       skipped rows, not from any real blank-row separator. Removed rather
       than shipped with a known false-positive; a safe fix needs
       `segment_blocks` to preserve the blank-vs-skipped distinction and
       hand it down, which is out of this deviation's scope (logged in
       deferred-items.md).
    3. **Marker rows with no open block** (no team-name row has appeared yet
       anywhere in the block): these form their own "headerless" group --
       `source_team1`/`source_team2` both `None`, no team identity derivable
       from the sheet alone. Surfaced via the `n Pair-Block-Gruppe(n) ohne
       Team-Namenspaar-Kopfzeile` notice in `segment_games` for the
       `docs/hc-blocks-ohne-kopfzeile.md` worksheet.

    Returns a list of `(rows, header_team1, header_team2)` tuples --
    `header_team1`/`header_team2` are the RAW values of the block-opening
    team-name row (`None`/`None` for a headerless marker-only group).

    A rule-1-only run (no marker rows anywhere in the block -- the real
    `Data`-tab pair block's only observed pattern, verified 2026-09-03
    against `data/raw/hc_files/Scoring Probability by Situation
    2023-2026.xlsx`: 0 O/D/S marker rows, 0 headerless groups) reproduces
    the validated 137 -> 22 unordered-pair collapse exactly (M3-02-
    RESEARCH.md Sec 1.2) -- confirmed by re-running this function against
    the real block this session, not assumed.

    This rule change invalidates every existing `block_key` in a pair block
    (fewer, larger games renumber the `game_index` sequence) --
    `data/reference/hc_games.csv` must be regenerated by re-running the
    ingest, never hand-patched (M3-02-RESEARCH.md Pitfall 2; see
    `docs/hc-workbook-ingest.md` Sec Wartung).
    """
    groups: list[tuple[list[tuple[int, tuple]], str | None, str | None]] = []
    current: list[tuple[int, tuple]] = []
    current_header: tuple[str, str] | None = None
    current_pair_key: frozenset[str] | None = None
    current_has_marker = False

    def _flush() -> None:
        if not current:
            return
        t1, t2 = current_header if current_header is not None else (None, None)
        groups.append((list(current), t1, t2))

    for row_num, values in rows:
        t1_raw = values[0] if len(values) >= 1 else None
        t2_raw = values[1] if len(values) >= 2 else None
        marker = _pair_row_marker(t1_raw, t2_raw)

        if marker is not None:
            current.append((row_num, values))
            current_has_marker = True
        else:
            pair_key = frozenset({_normalize_pair_label(t1_raw), _normalize_pair_label(t2_raw)})
            same_pair_continuation = (
                current
                and current_header is not None
                and not current_has_marker
                and pair_key == current_pair_key
            )
            if same_pair_continuation:
                current.append((row_num, values))
            else:
                if current:
                    _flush()
                current = [(row_num, values)]
                current_header = (
                    str(t1_raw) if t1_raw is not None else "",
                    str(t2_raw) if t2_raw is not None else "",
                )
                current_pair_key = pair_key
                current_has_marker = False

    _flush()
    return groups


def segment_games(block: HcBlock) -> tuple[list[HcGameSlice], list[str]]:
    """Split one `HcBlock` into per-game `HcGameSlice`s.

    A numeric block splits wherever `PLAY #` does not increase relative to
    the previous row (reset to 1, any decrease, or an unparseable cell -- see
    `_split_numeric_block`). A pair block splits per `_split_pair_block`'s
    combined rule: unordered `{team1, team2}` pair equality for a
    team-name-per-row stretch (possession flips do not split), or a
    team-name row opening a new block for the header + O/D/S-marker
    convention -- see `_split_pair_block`'s docstring for the full rule
    (Frage 2, confirmed 2026-09-03) and its documented, deliberate gap
    (a genuinely blank row cannot currently be distinguished from any other
    row `segment_blocks` silently skips).

    `block_key` is `b{block.index:02d}-g{game_index:02d}`, zero-padded and
    stable across runs for an unchanged sheet -- two blocks in the same sheet
    never collide (distinct `block_index`) and the same `block_key` string in
    two different sheets is disambiguated by the caller's
    `(workbook, sheet, block_key)` lookup, not by this function.

    A pair-block group with no team-name header (`source_team1 is None`) is
    "headerless" -- a run of O/D/S-marker rows with no preceding team-name
    row to inherit from anywhere earlier in the block. `segment_games`
    counts these and, when non-zero, emits one notice naming the count for
    `docs/hc-blocks-ohne-kopfzeile.md`.
    """
    messages: list[str] = []
    slices: list[HcGameSlice] = []

    if not block.rows:
        return slices, messages

    if block.kind == "numeric":
        header_groups: list[tuple[list[tuple[int, tuple]], str | None, str | None]] = [
            (group, None, None) for group in _split_numeric_block(block.rows, messages)
        ]
    else:
        header_groups = _split_pair_block(block.rows)

    n_headerless = 0
    for game_index, (group_rows, source_team1, source_team2) in enumerate(header_groups):
        first_row = group_rows[0][0]
        last_row = group_rows[-1][0]
        block_key = f"b{block.index:02d}-g{game_index:02d}"

        if block.kind == "pair" and source_team1 is None:
            n_headerless += 1

        slices.append(
            HcGameSlice(
                block_index=block.index,
                game_index=game_index,
                block_key=block_key,
                kind=block.kind,
                rows=group_rows,
                first_row=first_row,
                last_row=last_row,
                source_team1=source_team1,
                source_team2=source_team2,
            )
        )

    if n_headerless:
        messages.append(
            f"{n_headerless} Pair-Block-Gruppe(n) ohne Team-Namenspaar-Kopfzeile "
            "(O/D/S-Marker-Zeile(n) ohne vorausgehende Team-Namenszeile im selben "
            "Block) -- siehe docs/hc-blocks-ohne-kopfzeile.md"
        )

    return slices, messages


@dataclass(frozen=True)
class HcGameIdentity:
    """The resolved identity of one `HcGameSlice`: either a mapped row from
    `data/reference/hc_games.csv` (`provisional=False`) or a degraded
    placeholder (`provisional=True`) built when no row matches. Nothing is
    invented for a provisional identity -- `home_team`/`away_team`/`tier`/
    `season`/`game_date`/`corpus_game_id` all stay `None`.
    """

    game_id: str
    home_team: str | None
    away_team: str | None
    competition: str | None
    season: int | None
    game_date: str | None
    tier: str | None
    corpus_game_id: str | None
    provisional: bool


# HC-D04: `reference.map_teams` is deliberately NOT used for game identity --
# it raises `UnmappedTeamError` on any label with no mapping row, but an HC
# game the maintainer has not yet transcribed into `hc_games.csv` must
# degrade into a provisional id plus a loud notice, not abort the whole
# ingest run. Game identity for this source therefore comes from a direct,
# never-raising filter against the maintained `hc_games` frame below.
def resolve_game_identity(
    slice_: HcGameSlice,
    workbook_slug: str,
    sheet_slug: str,
    hc_games: pl.DataFrame,
) -> tuple[HcGameIdentity, list[str]]:
    """Resolve `slice_`'s game identity against the maintained `hc_games` frame.

    Filters on `(workbook, sheet, block_key)`. On a hit, returns the mapped
    identity built from that row (`provisional=False`), no message. On a
    miss, returns a provisional identity (`game_id =
    "hc-{workbook}-{sheet}-{block_key}"`, every other field `None`) plus
    exactly one notice naming the block key, the physical Excel row range,
    the play count and -- for a pair block -- the raw team labels, so a
    maintainer can add the `hc_games.csv` row without opening Excel. Never
    raises.
    """
    messages: list[str] = []

    match = hc_games.filter(
        (pl.col("workbook") == workbook_slug)
        & (pl.col("sheet") == sheet_slug)
        & (pl.col("block_key") == slice_.block_key)
    )

    if match.height:
        row = match.row(0, named=True)
        identity = HcGameIdentity(
            game_id=row["game_id"],
            home_team=row["home_team"],
            away_team=row["away_team"],
            competition=row["competition"],
            season=row["season"],
            game_date=row["game_date"],
            tier=row["tier"],
            corpus_game_id=row["corpus_game_id"],
            provisional=False,
        )
        return identity, messages

    provisional_id = f"hc-{workbook_slug}-{sheet_slug}-{slice_.block_key}"
    identity = HcGameIdentity(
        game_id=provisional_id,
        home_team=None,
        away_team=None,
        competition=None,
        season=None,
        game_date=None,
        tier=None,
        corpus_game_id=None,
        provisional=True,
    )

    n_plays = len(slice_.rows)
    team_note = ""
    if slice_.kind == "pair":
        team_note = f", Teams laut Sheet: {slice_.source_team1!r} vs. {slice_.source_team2!r}"
    messages.append(
        f"Unbekanntes Spiel {slice_.block_key!r} in {workbook_slug}/{sheet_slug}: "
        f"Zeilen {slice_.first_row}-{slice_.last_row} ({n_plays} Plays){team_note} -- "
        f"provisorische game_id {provisional_id!r} vergeben; data/reference/hc_games.csv "
        "um diese Zeile ergänzen"
    )

    return identity, messages


def _drop_placeholder_rows(game_df: pl.DataFrame) -> tuple[pl.DataFrame, int]:
    """Remove a numeric-block game's placeholder rows (`ODK`/`DN`/`RESULT`
    all null) before validation ever sees them -- HC corpus admission rule 1,
    confirmed 2026-09-04 (`.planning/todos/pending/
    2026-09-04-hc-korpus-zulassung-vor-training.md`).

    Typically the first `PLAY # 1-2` rows of a newly-started SP-charted
    game: nothing was charted there yet, so they are not real plays --
    previously they made the whole game FAIL `validation.checks.
    downs_range` (null `DN`) and get quarantined. A row with a real
    `RESULT` but null `DN` is a genuinely, if incompletely, charted play and
    stays -- checked by `downs_range` exactly as before. Classification
    only: this never weakens `downs_range` itself, it removes rows that
    were never real plays before the check runs. Returns `(df, n_removed)`.
    """
    if game_df.height == 0 or not {"ODK", "DN", "RESULT"}.issubset(game_df.columns):
        return game_df, 0
    n_before = game_df.height
    filtered = game_df.filter(
        ~(pl.col("ODK").is_null() & pl.col("DN").is_null() & pl.col("RESULT").is_null())
    )
    return filtered, n_before - filtered.height


def _reassign_hc_play_no(game_df: pl.DataFrame) -> tuple[pl.DataFrame, int]:
    """Reassign `PLAY #` from every row's 1-based position within the game --
    not just null cells -- HC corpus admission rule 2, confirmed 2026-09-04
    (`.planning/todos/pending/2026-09-04-hc-korpus-zulassung-vor-training.md`).

    The head coach's own `PLAY #` numbering (real charting gaps, resets, or
    nulls) previously fed straight into `play_id` via `hudl.
    derive_identity_columns`'s `PLAY #` -> `play_id` cast, so
    `validation.checks.gapless_play_ids` was checking HIS numbering, not the
    sheet's real row order -- a real charting gap (rows numbered e.g.
    1, 2, 5, 6) failed the check even though every row is a real,
    correctly-ordered play. The original value survives as `hc_play_no`
    before this overwrite; `conform_to_canonical` currently drops it (no
    slot for it in `canonical.CANONICAL_COLUMNS` yet -- extending that
    schema is out of this change's scope, `canonical.py` is a concurrent
    plan's territory), so it is not invented, only not yet persisted to
    `plays.parquet`.

    Returns `(df, n_rows)` -- `n_rows` is `game_df.height`: every row is
    reassigned now, unlike the previous null-only fill this replaces.
    """
    n_rows = game_df.height
    if n_rows == 0 or "PLAY #" not in game_df.columns:
        return game_df, n_rows

    game_df = game_df.with_columns(pl.col("PLAY #").alias("hc_play_no"))
    game_df = game_df.with_row_index(name="_hc_pos", offset=1)
    game_df = game_df.with_columns(pl.col("_hc_pos").cast(pl.Utf8).alias("PLAY #"))
    game_df = game_df.drop("_hc_pos")
    return game_df, n_rows


# HC corpus admission rule 3 (confirmed 2026-09-04): a provisional game (no
# hc_games.csv row, so `home_team`/`away_team` both resolve to `None`) still
# derives `posteam`/`defteam` from a real `ODK` value, using these
# game-scoped placeholder labels -- "Teamnamen sind nur für Dedupe/Splits
# relevant", EP only needs the offense/defense perspective. The real
# `home_team`/`away_team` COLUMNS (stamped from `identity.home_team`/
# `identity.away_team` in `ingest_workbook`, not by this function) stay null
# for a provisional game exactly as before -- only `posteam`/`defteam` get a
# fallback, so `hc_dedupe`'s fingerprint comparison (which reads
# `home_team`/`away_team`, never `posteam`/`defteam`) is unaffected.
_PROVISIONAL_POSTEAM_PLACEHOLDER = "HC-OFF"
_PROVISIONAL_DEFTEAM_PLACEHOLDER = "HC-DEF"


def _stamp_posteam_defteam(
    game_df: pl.DataFrame, kind: str, home_team: str | None, away_team: str | None
) -> pl.DataFrame:
    """Derive `posteam`/`defteam` for one game's rows, from `ODK` alone --
    not from `kind`.

    A `numeric` block follows hudl's own convention: `posteam = home_team`
    when `ODK == "O"`, `away_team` for every other non-null `ODK` (`"D"`,
    `"K"` or `"S"` alike -- the `"K"`/`"S"` overrides only change
    `play_type`, not `posteam`), null when `ODK` itself is null. A `pair`
    block's team-name header rows keep `ODK` null (`_null_pair_block_tail`),
    so they fall into the same null branch by construction -- the offence
    side stays undetermined for those rows, as before. A `pair` block's
    O/D/S-marker rows (Frage 2, confirmed 2026-09-03) DO carry a real `ODK`
    value now, so this same rule derives their `posteam`/`defteam` from the
    slice's resolved `home_team`/`away_team` too -- `kind` is kept as a
    parameter only for callers/readability, it no longer branches behaviour.

    When `home_team`/`away_team` are both `None` (a provisional game, HC
    corpus admission rule 3), `_PROVISIONAL_POSTEAM_PLACEHOLDER`/
    `_PROVISIONAL_DEFTEAM_PLACEHOLDER` stand in instead, so a real `ODK`
    value still resolves a usable offense/defense perspective rather than
    nulling out for want of a declared team name.
    """
    effective_home = home_team if home_team is not None else _PROVISIONAL_POSTEAM_PLACEHOLDER
    effective_away = away_team if away_team is not None else _PROVISIONAL_DEFTEAM_PLACEHOLDER
    game_df = game_df.with_columns(
        pl.when(pl.col("ODK") == "O")
        .then(pl.lit(effective_home, dtype=pl.Utf8))
        .when(pl.col("ODK").is_not_null())
        .then(pl.lit(effective_away, dtype=pl.Utf8))
        .otherwise(pl.lit(None, dtype=pl.Utf8))
        .alias("posteam")
    )
    game_df = game_df.with_columns(
        pl.when(pl.col("posteam") == pl.lit(effective_home, dtype=pl.Utf8))
        .then(pl.lit(effective_away, dtype=pl.Utf8))
        .when(pl.col("posteam").is_not_null())
        .then(pl.lit(effective_home, dtype=pl.Utf8))
        .otherwise(pl.lit(None, dtype=pl.Utf8))
        .alias("defteam")
    )
    return game_df


def count_result_tokens(df: pl.DataFrame) -> dict[str, int]:
    """Count every token observed in `RESULT` (split on `", "`), independent
    of the contract's known vocabulary -- so a caller can report per-token
    counts (including any not-yet-ratified token, e.g. an HC-only one, see
    docs/hc-rueckfragen-2026-09.md Frage 3) without re-parsing RESULT itself.
    """
    if df.height == 0 or "RESULT" not in df.columns:
        return {}

    tokens = (
        df["RESULT"]
        .fill_null("")
        .str.split(", ")
        .list.eval(pl.element().filter(pl.element() != ""))
        .explode()
        .drop_nulls()
        .to_list()
    )
    counts: dict[str, int] = {}
    for token in tokens:
        counts[token] = counts.get(token, 0) + 1
    return counts


def ingest_workbook(
    path: Path,
    sheet: str,
    contract: Contract,
    hc_games: pl.DataFrame,
    player_mapping: pl.DataFrame,
) -> tuple[pl.DataFrame, HcIngestNotices]:
    """Turn one (workbook, sheet)'s charted rows into canonical plays.

    Order of operations, mirroring `hudl.ingest_file` and reusing its
    derivation functions unchanged (HC-D01): `read_sheet_rows` ->
    `segment_blocks` -> per block `map_block_to_frame` -> `segment_games` ->
    per game slice `resolve_game_identity`, `_drop_placeholder_rows`
    (numeric blocks only, HC corpus admission rule 1), constant-stamping
    (`source`, `competition`, `season`, `game_id`, `game_date`, `home_team`,
    `away_team`, `half`, `result_raw`), `_reassign_hc_play_no` (HC corpus
    admission rule 2), `posteam`/`defteam` (HC corpus admission rule 3 for a
    provisional game) -> concatenate the sheet's game frames (`how="vertical"`; they
    share a schema by construction -- a mismatch is a bug, left to raise
    rather than silently switched to `diagonal`) -> `hudl.derive_identity_columns`
    -> `hudl.parse_result_tokens` -> `hudl.derive_outcome_columns` -> ODK
    `"K"` kickoff override (identical to hudl's) -> pair-block ODK `"S"`
    no-play override (Frage 2, Antwort 2026-09-03) -> `hudl.derive_drive_id` ->
    `add_scoring_play_team(credit_defense=True)` -> `add_score_columns` ->
    `hudl.derive_yards_gained_first_down` -> `reference.map_players` (source
    key `"hc_workbook"`, deliberately coarse -- one `player_mapping.csv` row
    serves all three workbooks, not a per-sheet key) -> `count_result_tokens`
    -> `conform_to_canonical`.

    `half` rule (M3-02-RESEARCH.md Sec 2.2): HC workbooks carry no real
    half-boundary data, so `half` is not derived -- it is a per-game constant
    `HALF_SENTINEL` (`2`) for a game declared in `data/reference/hc_games.csv`
    and charted on a sheet outside `HALF_SENTINEL_EXCLUDED_SHEETS`, else
    `null`. `2` is the only sentinel that both satisfies
    `validation.checks.half_assigned` honestly and keeps
    `features.mutations._mark_half_end`'s `game_end` (WP `Winner`, EP's
    post-game `None`-ing) firing correctly; the named cost is that no real
    halftime `No_Score` boundary exists for these games, so a scoreless
    first-half drive is backward-filled with the game's next actual score.

    Never raises on a data-quality finding -- everything folds into the
    returned `HcIngestNotices`. A sheet with no usable blocks at all (e.g.
    entirely empty) returns a zero-row `CANONICAL_COLUMNS` frame plus a
    notice, not an error. Structural problems (an absent sheet) still
    propagate from `read_sheet_rows` (`SheetNotFoundError`).
    """
    workbook_slug = slugify(Path(path).stem)
    sheet_slug = slugify(sheet)
    source_label = hc_source_label(path, sheet)

    header, rows, messages = read_sheet_rows(path, sheet)
    blocks, block_messages = segment_blocks(header, rows)
    messages.extend(block_messages)

    game_frames: list[pl.DataFrame] = []
    header_reports: list[HeaderReport] = []
    domain_violations: list[DomainViolation] = []
    pair_row_total = 0
    pair_marker_row_total = 0
    synthesized_play_ids = 0
    n_sentinel_rows = 0
    n_sentinel_games = 0
    n_null_undeclared_rows = 0
    n_null_copy_of_data_rows = 0
    placeholder_row_total = 0
    placeholder_game_count = 0

    for block in blocks:
        block_df, header_report, block_domain, block_map_messages = map_block_to_frame(
            block, contract
        )
        messages.extend(block_map_messages)
        header_reports.append(header_report)
        domain_violations.extend(block_domain)

        slices, segment_messages = segment_games(block)
        messages.extend(segment_messages)

        offset = 0
        for slice_ in slices:
            n = len(slice_.rows)
            game_df = block_df.slice(offset, n)
            offset += n

            # A pair block's frame carries the two synthetic hc_pair_team1/
            # hc_pair_team2 columns (_null_pair_block_tail); a numeric
            # block's frame never does. Drop them here -- the raw labels
            # already live in slice_.source_team1/source_team2 for identity
            # resolution, and keeping them would break the "every game frame
            # in this sheet shares one schema" concat invariant below.
            present_synthetic = [c for c in _SYNTHETIC_COLUMNS if c in game_df.columns]
            if present_synthetic:
                game_df = game_df.drop(present_synthetic)

            identity, identity_messages = resolve_game_identity(
                slice_, workbook_slug, sheet_slug, hc_games
            )
            messages.extend(identity_messages)

            # HC corpus admission rule 1 (confirmed 2026-09-04): a numeric
            # block's placeholder rows (ODK/DN/RESULT all null -- typically
            # PLAY # 1-2 at a new SP-charted game's start) are removed
            # before validation ever runs. Scoped to numeric blocks only: a
            # pair block's rows never carry a real ODK/DN pair this test can
            # evaluate (`_null_pair_block_tail` nulls both), and the
            # `row_markers` series built below from `slice_.rows`
            # (unfiltered) would desync against a filtered pair `game_df`.
            if slice_.kind == "numeric":
                game_df, n_placeholder = _drop_placeholder_rows(game_df)
                if n_placeholder:
                    placeholder_row_total += n_placeholder
                    placeholder_game_count += 1
                    messages.append(
                        f"{n_placeholder} Platzhalter-Zeile(n) (ODK/DN/RESULT alle "
                        f"null) vor der Validierung entfernt: {workbook_slug}/"
                        f"{sheet_slug} game {identity.game_id!r}"
                    )
                n = game_df.height

            # half=2 sentinel (M3-02-RESEARCH.md Sec 2.2), stamped alongside
            # the other per-game identity columns rather than blanket-null
            # after the concat: it is a per-game decision (declared vs.
            # undeclared), never a per-sheet one. `Copy of Data`'s sheet
            # exclusion overrides the declaration outright (Frage 2, Sec 1.3)
            # -- undeclared and never-half-assigned rows keep quarantining.
            use_sentinel = (not identity.provisional) and sheet not in HALF_SENTINEL_EXCLUDED_SHEETS
            game_df = game_df.with_columns(
                pl.lit(source_label, dtype=pl.Utf8).alias("source"),
                pl.lit(identity.competition, dtype=pl.Utf8).alias("competition"),
                pl.lit(identity.season, dtype=pl.Int32).alias("season"),
                pl.lit(identity.game_id, dtype=pl.Utf8).alias("game_id"),
                pl.lit(identity.game_date, dtype=pl.Utf8).alias("game_date"),
                pl.lit(identity.home_team, dtype=pl.Utf8).alias("home_team"),
                pl.lit(identity.away_team, dtype=pl.Utf8).alias("away_team"),
                pl.lit(HALF_SENTINEL if use_sentinel else None, dtype=pl.Int32).alias("half"),
                pl.col("RESULT").alias("result_raw"),
            )
            if use_sentinel:
                n_sentinel_rows += n
                n_sentinel_games += 1
            elif sheet in HALF_SENTINEL_EXCLUDED_SHEETS:
                n_null_copy_of_data_rows += n
            else:
                n_null_undeclared_rows += n

            game_df, n_reassigned = _reassign_hc_play_no(game_df)
            synthesized_play_ids += n_reassigned

            game_df = _stamp_posteam_defteam(
                game_df, slice_.kind, identity.home_team, identity.away_team
            )
            if slice_.kind == "pair":
                pair_row_total += n
                # Per-row marker value ("O"/"D"/"S"/None), carried through
                # the concat so the post-concat ODK == "S" -> "no_play"
                # override below only ever touches rows that actually came
                # from a pair-block marker row -- never a genuine numeric-
                # block ODK == "S" row (371 real occurrences in the corpus,
                # unrelated pre-existing behaviour, out of this plan's
                # scope). Dropped again before conform_to_canonical.
                row_markers = [
                    _pair_row_marker(
                        v[0] if len(v) >= 1 else None, v[1] if len(v) >= 2 else None
                    )
                    for _, v in slice_.rows
                ]
                pair_marker_row_total += sum(1 for m in row_markers if m is not None)
                game_df = game_df.with_columns(
                    pl.Series("_hc_pair_marker_odk", row_markers, dtype=pl.Utf8)
                )
            else:
                game_df = game_df.with_columns(
                    pl.lit(None, dtype=pl.Utf8).alias("_hc_pair_marker_odk")
                )

            game_frames.append(game_df)

    if placeholder_row_total:
        messages.append(
            f"{placeholder_row_total} Platzhalter-Zeile(n) insgesamt in "
            f"{placeholder_game_count} Spiel(en) vor der Validierung entfernt "
            "(ODK/DN/RESULT alle null, HC-Korpus-Zulassungsregel 1, 2026-09-04)"
        )
    if pair_row_total:
        pair_header_row_total = pair_row_total - pair_marker_row_total
        messages.append(
            f"{pair_row_total} Pair-Block-Zeile(n) insgesamt: {pair_marker_row_total} "
            "Zeile(n) mit O/D/S-Marker (ODK/posteam/defteam aus dem Marker abgeleitet, "
            f"Frage 2 Antwort 2026-09-03), {pair_header_row_total} Zeile(n) mit "
            "Team-Namenspaar-Kopfzeile statt echtem ODK (posteam/defteam bleiben null; "
            "der Spaltenversatz ab RECEIVED BY bleibt für beide Zeilenarten offen)"
        )
    if synthesized_play_ids:
        messages.append(
            f"PLAY # für {synthesized_play_ids} Zeile(n) aus der Zeilenreihenfolge "
            "innerhalb des jeweiligen Spiels neu vergeben (hc_play_no behält die "
            "Original-PLAY #, sofern im Sheet vorhanden -- HC-Korpus-Zulassungsregel 2, "
            "2026-09-04: gapless_play_ids prüft damit die reale Reihenfolge, nicht "
            "seine Nummerierung)"
        )

    if not game_frames:
        df = pl.DataFrame(schema={**CORE_COLUMNS, **NULLABLE_EXTRAS})
        messages.append(
            f"Sheet {sheet!r} enthält keine auswertbaren Blöcke: leerer kanonischer "
            "Frame zurückgegeben"
        )
    else:
        df = pl.concat(game_frames, how="vertical")

        df = hudl.derive_identity_columns(df)
        df = hudl.parse_result_tokens(df)
        df, outcome_messages = hudl.derive_outcome_columns(df)
        messages.extend(outcome_messages)

        # ODK == 'K' overrides play_type to "kickoff" regardless of RESULT
        # tokens, identical to hudl's own override -- harmless when no 'K'
        # row exists in this sheet.
        df = df.with_columns(
            pl.when(pl.col("ODK") == "K")
            .then(pl.lit("kickoff"))
            .otherwise(pl.col("play_type"))
            .alias("play_type")
        )

        # Pair-block O/D/S marker "S" -> no_play (Frage 2, Antwort
        # 2026-09-03: "S für no-play"), handled like Timeout/Offsetting
        # Penalties (contract v1.2). Scoped to `_hc_pair_marker_odk` (set
        # only for pair-block marker rows, see the per-slice loop above) so
        # this NEVER touches a genuine numeric-block ODK == "S" row -- that
        # is pre-existing, already-shipped behaviour, out of this plan's
        # scope.
        n_pair_marker_s = int((df["_hc_pair_marker_odk"] == "S").sum())
        if n_pair_marker_s:
            df = df.with_columns(
                pl.when(pl.col("_hc_pair_marker_odk") == "S")
                .then(pl.lit("no_play"))
                .otherwise(pl.col("play_type"))
                .alias("play_type")
            )
            messages.append(
                f"{n_pair_marker_s} Pair-Block-Zeile(n) mit O/D/S-Marker 'S' als "
                "play_type='no_play' markiert (kein echter Play, Frage 2 Antwort "
                "2026-09-03)"
            )
        df = df.drop("_hc_pair_marker_odk")

        df = hudl.derive_drive_id(df)

        # half=2 sentinel notices (M3-02-RESEARCH.md Sec 2.2): built from the
        # per-game counts accumulated above, each emitted only when non-zero,
        # so a sheet with only declared or only undeclared games never
        # reports a spurious zero count.
        if n_sentinel_rows:
            messages.append(
                f"half = {HALF_SENTINEL} (Sentinel) für {n_sentinel_rows} Zeile(n) aus "
                f"{n_sentinel_games} in hc_games.csv deklarierten Spiel(en) gesetzt -- "
                "HC-Workbooks tragen keine echte Halbzeitgrenze; Folge: kein "
                "No_Score-Marker zur Halbzeit, eine torlose Drive der ersten Halbzeit "
                "erbt den nächsten tatsächlichen Score"
            )
        n_null_half_rows = n_null_undeclared_rows + n_null_copy_of_data_rows
        if n_null_half_rows:
            messages.append(
                f"half = null für {n_null_half_rows} Zeile(n) ({n_null_undeclared_rows} aus "
                "nicht in hc_games.csv deklarierten Spielen, "
                f"{n_null_copy_of_data_rows} aus 'Copy of Data' -- Frage 2 offen): bleiben "
                "in Quarantäne (half_assigned)"
            )

        df = add_scoring_play_team(df, credit_defense=True)
        df = add_score_columns(df)
        df = hudl.derive_yards_gained_first_down(df)

    # Coarse "hc_workbook" source key (not the per-sheet source_label): one
    # player_mapping.csv row then serves all three HC workbooks, since the
    # HC's own player-label vocabulary does not vary per workbook/sheet.
    player_columns = ["qb", "thrown_by", "received_by", "target", "tackle"]
    map_result = map_players(df, player_mapping, source="hc_workbook", columns=player_columns)
    df = map_result.frame
    unmapped_players = map_result.unmapped
    if unmapped_players:
        # PII discipline: only the count, never the labels themselves.
        messages.append(
            f"{len(unmapped_players)} nicht zugeordnete Spieler-Label in "
            f"{player_columns}"
        )

    result_token_counts = count_result_tokens(df)

    df, conform_report = conform_to_canonical(df, source_label)

    notices = HcIngestNotices(
        source_label=source_label,
        sheet=sheet,
        messages=messages,
        header=header_reports[0] if header_reports else None,
        domain=domain_violations,
        conform=conform_report,
        unmapped_players=unmapped_players,
        result_token_counts=result_token_counts,
    )
    return df, notices
