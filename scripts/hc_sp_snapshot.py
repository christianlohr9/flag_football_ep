"""Read-only snapshot of the head coach's aggregate SP/EP-by-situation tabs.

Takes the nine aggregate analysis tabs in `Scoring Probability by Situation
2023-2026.xlsx` (gitignored PII workbook under `data/raw/hc_files/`) and
writes them out as tidy, deterministic CSVs under `data/reference/hc_sp_tables/`
so the head coach's own method can be reproduced and quoted (M3-02 EPA-D03)
without ever re-opening the workbook. Per M3-02-RESEARCH section 4, all nine
tabs contain only down, distance, field-half, probability, count and
regression-coefficient values -- no player name, no team name, no real date.

Two things this script does beyond a straight copy:

- `Sample Size by D&D` is snapshotted alongside every probability tab, key for
  key, so a point estimate on n=13 is never presented without the n sitting
  next to it.
- Three of the `Clustered` tabs' distance-bin row labels were silently
  autocorrected by Excel into `datetime` objects (a well-known `MM-DD` ->
  date trap, RESEARCH Pitfall 4, section 4.3). `reconstruct_distance_bin`
  turns them back into text labels (`1-5`, `6-10`, `11-15`), tagged
  `distance_bin_source == "reconstructed"` in the output, never silently
  presented as read facts.

The `Data` and `Copy of Data` tabs -- the only two tabs in this workbook that
carry player labels -- are named in `FORBIDDEN_SHEETS` and are unreachable by
construction: `_load_sheet` raises before any such sheet could be opened.
"""

from __future__ import annotations

import argparse
import datetime
import hashlib
import sys
from pathlib import Path
from typing import Any

import openpyxl
import polars as pl

REPO_ROOT = Path(__file__).resolve().parent.parent

DEFAULT_WORKBOOK = REPO_ROOT / "data" / "raw" / "hc_files" / "Scoring Probability by Situation 2023-2026.xlsx"
DEFAULT_OUT_DIR = REPO_ROOT / "data" / "reference" / "hc_sp_tables"

# The nine aggregate tabs this script is allowed to open. Anything else in
# the workbook -- most importantly `Data` / `Copy of Data`, which carry
# player labels -- must never be reachable through `_load_sheet`.
ALLOWED_SHEETS: tuple[str, ...] = (
    "SP by D&D",
    "EP by D&D",
    "Sample Size by D&D",
    "SP by D&D Clustered",
    "EP by D&D Clustered",
    "Sample Size by D&D Clustered",
    "SP by D&D weighted",
    "EP by D&D weighted",
    "Reg",
)

FORBIDDEN_SHEETS: tuple[str, ...] = ("Data", "Copy of Data")

# tab name -> output CSV filename (matches this plan's `files_modified`)
TAB_OUTPUT_FILES: dict[str, str] = {
    "SP by D&D": "sp_by_dd.csv",
    "EP by D&D": "ep_by_dd.csv",
    "Sample Size by D&D": "sample_size_by_dd.csv",
    "SP by D&D Clustered": "sp_by_dd_clustered.csv",
    "EP by D&D Clustered": "ep_by_dd_clustered.csv",
    "Sample Size by D&D Clustered": "sample_size_by_dd_clustered.csv",
    "SP by D&D weighted": "sp_by_dd_weighted.csv",
    "EP by D&D weighted": "ep_by_dd_weighted.csv",
}

MATRIX_TABS: tuple[str, ...] = tuple(TAB_OUTPUT_FILES.keys())

# Half-header text -> normalised `field_half` value. Any other non-blank
# string encountered in the half-header row is a corruption/typo in the
# workbook this script has never seen before -- raise rather than default.
_HALF_HEADER_MAP: dict[str, str] = {
    "Own Half": "own",
    "Opposite Half": "opponent",
}

# Row-0 labels that mark a grand-total / summary row rather than a real
# (down, distance, field_half) cell. Excel-typo tolerant: compared
# case-insensitively after stripping whitespace.
_SUMMARY_ROW_LABELS = {"gesamt", "total", "summe"}

# PII / domain guard from this plan's <pii_discipline> block: no output cell
# may be a string longer than this, and no output column may hold a
# `datetime` (checked separately in `_assert_no_datetime_or_long_strings`).
_MAX_CELL_STRING_LEN = 24


class ForbiddenSheetError(ValueError):
    """Raised when code asks to read a sheet not in `ALLOWED_SHEETS`."""


def _load_sheet(wb: openpyxl.Workbook, name: str) -> Any:
    """Return `wb[name]`, refusing anything not in `ALLOWED_SHEETS`.

    This is the single choke point that keeps `Data` / `Copy of Data` -- the
    two player-labelled tabs -- unreachable even by a later edit that adds a
    new call site elsewhere in this file.
    """
    if name in FORBIDDEN_SHEETS:
        raise ForbiddenSheetError(
            f"refusing to open {name!r}: forbidden sheet (carries player labels)"
        )
    if name not in ALLOWED_SHEETS:
        raise ForbiddenSheetError(
            f"refusing to open {name!r}: not in ALLOWED_SHEETS {ALLOWED_SHEETS}"
        )
    return wb[name]


def reconstruct_distance_bin(value: Any) -> tuple[str, str]:
    """Return `(label, provenance)` for one distance-bin cell.

    `provenance` is `"read"` when `value` is already text or a plain number.
    For a `datetime.datetime` -- the Excel-autocorrect corruption documented
    in M3-02-RESEARCH section 4.3 (Pitfall 4: `"1-5"` -> `date(2021, 1, 5)`,
    `"6-10"` -> `date(2021, 6, 10)`, `"11-15"` -> `date(2021, 11, 15)`) --
    `provenance` is `"reconstructed"` and the label is rebuilt from the
    month/day as `f"{month}-{day}"`.

    Raises `ValueError` if a `datetime`'s year is not 2021: that would mean
    the corruption pattern differs from the one documented this session, and
    blindly reconstructing it would risk fabricating a wrong label.
    """
    if isinstance(value, datetime.datetime):
        if value.year != 2021:
            raise ValueError(
                f"datetime cell {value!r} has year != 2021 -- corruption pattern "
                "does not match M3-02-RESEARCH section 4.3; refusing to guess a "
                "reconstruction"
            )
        return f"{value.month}-{value.day}", "reconstructed"
    if isinstance(value, (int, float)):
        label = str(int(value)) if float(value).is_integer() else str(value)
        return label, "read"
    label = str(value).strip()
    return label, "read"


def _column_half_map(header_half_row: tuple[Any, ...]) -> dict[int, str]:
    """Forward-fill the half-header row into a per-column `field_half` map.

    Only string cells update the "current half"; a stray leading numeric
    cell (seen in the `weighted` tabs' row 0, column 0) or a blank separator
    column simply doesn't change it. Columns to the left of the first
    recognised half label get no entry at all (excluded from down-column
    detection downstream) -- this is what keeps a spurious `1.0` in column 0
    of the `weighted` tabs from ever being mistaken for a down column.
    """
    column_half: dict[int, str] = {}
    current: str | None = None
    for col_idx, cell in enumerate(header_half_row):
        if isinstance(cell, str) and cell.strip():
            text = cell.strip()
            if text not in _HALF_HEADER_MAP:
                raise ValueError(
                    f"unrecognized field-half header {text!r} at column {col_idx}; "
                    f"expected one of {sorted(_HALF_HEADER_MAP)}"
                )
            current = _HALF_HEADER_MAP[text]
        if current is not None:
            column_half[col_idx] = current
    return column_half


def _down_columns(
    header_down_row: tuple[Any, ...], column_half: dict[int, str]
) -> dict[int, tuple[int, str]]:
    """Map column index -> `(down, field_half)` for every real down column.

    A column counts as a down column only if it (a) already has a
    `field_half` assigned via `_column_half_map` and (b) its row-1 cell is a
    whole number in `{1, 2, 3, 4}`. This excludes `Total` columns (text, not
    a down number) and the `weighted` tabs' spurious column-0 `1.0` (no
    `field_half` assigned, since column 0 precedes the first half label).
    """
    down_cols: dict[int, tuple[int, str]] = {}
    for col_idx, cell in enumerate(header_down_row):
        half = column_half.get(col_idx)
        if half is None:
            continue
        if isinstance(cell, bool) or not isinstance(cell, (int, float)):
            continue
        if not float(cell).is_integer():
            continue
        down = int(cell)
        if down in (1, 2, 3, 4):
            down_cols[col_idx] = (down, half)
    return down_cols


def read_matrix_tab(ws: Any) -> list[tuple[int, str, str, str, float]]:
    """Tidy `(down, distance_bin, distance_bin_source, field_half, value)`
    records from one down x (distance x field_half) matrix tab.

    Row 0 is the half-header (`Own Half` / `Opposite Half`, forward-filled
    across the merged-looking cells); row 1 is the down-number header. Data
    starts at row 2 (0-indexed row 2, i.e. spreadsheet row 3).

    Each field_half block gets its own distance-bin label column: the
    column immediately preceding that half's first down column. In the
    unclustered/weighted tabs that column is always blank in data rows, so
    the label falls back to column 0 (one shared distance-to-go value for
    both halves). In the `Clustered` tabs that column genuinely holds a
    second, slightly different bin-edge label per half (RESEARCH section
    4.3) -- so it is used directly when present.

    Row skip rules (never invent a cell):
    - A row whose column-0 label is `Gesamt`/`Total`/`Summe` (case-
      insensitive) is a grand-total summary row, not a per-bin cell -- skip
      entirely.
    - A row whose column-0 label is blank (`None`) but which still carries
      down-column values is a total-without-a-label artifact (seen in
      `Sample Size by D&D`) -- skip entirely, there is no distance to
      attach the value to.
    - A row where every down-column value is `None` (Excel's autofilled
      distance labels run well past where real data stops) is skipped.
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []

    header_half_row = rows[0]
    header_down_row = rows[1]
    column_half = _column_half_map(header_half_row)
    down_cols = _down_columns(header_down_row, column_half)

    # label column per half = (min down-column index for that half) - 1
    label_col_by_half: dict[str, int] = {}
    for half in set(h for _d, h in down_cols.values()):
        min_col = min(c for c, (_d, h) in down_cols.items() if h == half)
        label_col_by_half[half] = min_col - 1

    records: list[tuple[int, str, str, str, float]] = []
    for row in rows[2:]:
        primary_label = row[0]
        if isinstance(primary_label, str) and primary_label.strip().lower() in _SUMMARY_ROW_LABELS:
            continue
        row_values = {c: row[c] for c in down_cols}
        if all(v is None for v in row_values.values()):
            continue
        if primary_label is None and any(v is not None for v in row_values.values()):
            # total-without-a-label artifact (Sample Size by D&D row 51) --
            # no distance to attach the value to, skip.
            continue

        for col_idx, (down, half) in down_cols.items():
            value = row[col_idx]
            if value is None:
                continue
            label_col = label_col_by_half[half]
            raw_label = row[label_col] if label_col >= 0 else None
            if raw_label is None:
                raw_label = primary_label
            distance_bin, source = reconstruct_distance_bin(raw_label)
            records.append((down, distance_bin, source, half, float(value)))

    return records


def read_reg_tab(
    ws_values: Any, ws_formulas: Any
) -> list[tuple[str, int, str, str, str]]:
    """`(field_half, down, column, formula_text, kind)` records from `Reg`.

    `Reg` holds one row per (field_half, down, distance) cell, but the
    formula text is constant within a (field_half, down) block -- only the
    row-specific cell references change (e.g. `C2` vs `C50`). This function
    returns ONE record per (field_half, down, column) combination, taken
    verbatim from that block's first row, rather than one per distance row.

    `field_half` is read off the `Code` column's sign (own-half rows carry a
    leading `-`, e.g. `-11`; opposite-half rows carry the bare digits, e.g.
    `11`) -- the `Half` text column itself is blank/uninformative in the
    real workbook. Rows where `Down` is not in `{1, 2, 3, 4}` (two stray
    rows referencing `'General Stats'!G3`/`G4` at the tail of the sheet, not
    part of the down x distance grid) are excluded.

    `column` is `"Forecast"` (column F, `kind="forecast"`, a live
    `=FORECAST(...)` call) or `"Reg"` (column H, `kind="polynomial"`, a
    hardcoded polynomial with literal coefficients).
    """
    value_rows = list(ws_values.iter_rows(values_only=True))
    formula_rows = list(ws_formulas.iter_rows(values_only=True))

    seen: set[tuple[str, int]] = set()
    records: list[tuple[str, int, str, str, str]] = []

    for value_row, formula_row in zip(value_rows[1:], formula_rows[1:]):
        down_raw = value_row[1]
        code = value_row[3]
        if not isinstance(down_raw, (int, float)) or isinstance(down_raw, bool):
            continue
        if not float(down_raw).is_integer() or int(down_raw) not in (1, 2, 3, 4):
            continue
        down = int(down_raw)
        if not isinstance(code, str) or not code:
            continue
        half = "own" if code.startswith("-") else "opponent"

        key = (half, down)
        if key in seen:
            continue
        seen.add(key)

        forecast_formula = formula_row[5]
        polynomial_formula = formula_row[7]
        if isinstance(forecast_formula, str) and forecast_formula.startswith("="):
            records.append((half, down, "Forecast", forecast_formula, "forecast"))
        if isinstance(polynomial_formula, str) and polynomial_formula.startswith("="):
            records.append((half, down, "Reg", polynomial_formula, "polynomial"))

    return records


def _assert_no_datetime_or_long_strings(df: pl.DataFrame, *, tab: str) -> None:
    """Pre-write PII/domain assertions from this plan's `<pii_discipline>`
    block: no output cell is a string longer than 24 characters, no output
    column holds a `datetime`.
    """
    for col_name, dtype in zip(df.columns, df.dtypes):
        if dtype in (pl.Datetime, pl.Date):
            raise ValueError(
                f"{tab}: column {col_name!r} has dtype {dtype} -- a corrupted "
                "distance-bin label was not reconstructed to text before writing"
            )
        if dtype == pl.Utf8:
            too_long = df.filter(pl.col(col_name).str.len_chars() > _MAX_CELL_STRING_LEN)
            if too_long.height > 0:
                offending = too_long[col_name].to_list()[0]
                raise ValueError(
                    f"{tab}: column {col_name!r} has a string longer than "
                    f"{_MAX_CELL_STRING_LEN} characters: {offending!r}"
                )


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help=f"path to the HC SP-by-situation workbook (default: {DEFAULT_WORKBOOK})",
    )
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=DEFAULT_OUT_DIR,
        help=f"directory to write CSVs into (default: {DEFAULT_OUT_DIR})",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="print each tab's record count and reconstructed labels, write nothing",
    )
    args = parser.parse_args(argv)

    workbook_path: Path = args.workbook
    if not workbook_path.exists():
        print(f"workbook not found: {workbook_path}", file=sys.stderr)
        return 1

    wb_values = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    wb_formulas = openpyxl.load_workbook(workbook_path, data_only=False, read_only=True)
    try:
        tab_frames: dict[str, pl.DataFrame] = {}
        tab_reconstructed: dict[str, list[str]] = {}

        for tab in MATRIX_TABS:
            ws = _load_sheet(wb_values, tab)
            records = read_matrix_tab(ws)
            value_col = "n" if "Sample Size" in tab else "value"
            df = pl.DataFrame(
                records,
                schema=["down", "distance_bin", "distance_bin_source", "field_half", value_col],
                orient="row",
            )
            if value_col == "n":
                df = df.with_columns(pl.col("n").round(0).cast(pl.Int64))
            df = df.sort(["field_half", "down", "distance_bin"])
            tab_frames[tab] = df
            tab_reconstructed[tab] = sorted(
                set(
                    df.filter(pl.col("distance_bin_source") == "reconstructed")[
                        "distance_bin"
                    ].to_list()
                )
            )

        ws_reg_values = _load_sheet(wb_values, "Reg")
        ws_reg_formulas = _load_sheet(wb_formulas, "Reg")
        reg_records = read_reg_tab(ws_reg_values, ws_reg_formulas)
        reg_df = pl.DataFrame(
            reg_records,
            schema=["field_half", "down", "column", "formula_text", "kind"],
            orient="row",
        ).sort(["field_half", "down", "column"])

        if args.dry_run:
            total_reconstructed: set[str] = set()
            for tab in MATRIX_TABS:
                df = tab_frames[tab]
                print(f"[dry-run] {tab}: {df.height} records")
                if tab_reconstructed[tab]:
                    total_reconstructed.update(tab_reconstructed[tab])
            print(f"[dry-run] Reg: {reg_df.height} records")
            print(
                f"[dry-run] {len(total_reconstructed)} reconstructed bin labels: "
                f"{sorted(total_reconstructed)}"
            )
            print("[dry-run] nothing written")
            return 0

        args.out_dir.mkdir(parents=True, exist_ok=True)

        manifest_rows: list[tuple[str, str, str, int, int, str, str]] = []
        source_basename = workbook_path.name
        workbook_sha256 = _sha256(workbook_path)
        extracted_at = datetime.datetime.now(datetime.timezone.utc).date().isoformat()

        for tab in MATRIX_TABS:
            df = tab_frames[tab]
            _assert_no_datetime_or_long_strings(df, tab=tab)
            out_file = TAB_OUTPUT_FILES[tab]
            out_path = args.out_dir / out_file
            df.write_csv(out_path)
            n_reconstructed = len(tab_reconstructed[tab])
            manifest_rows.append(
                (tab, source_basename, workbook_sha256, df.height, n_reconstructed, extracted_at, out_file)
            )
            print(f"wrote {out_path} ({df.height} records)")

        _assert_no_datetime_or_long_strings(reg_df, tab="Reg")
        reg_out_path = args.out_dir / "reg_formulas.csv"
        reg_df.write_csv(reg_out_path)
        manifest_rows.append(
            ("Reg", source_basename, workbook_sha256, reg_df.height, 0, extracted_at, "reg_formulas.csv")
        )
        print(f"wrote {reg_out_path} ({reg_df.height} records)")

        manifest_df = pl.DataFrame(
            manifest_rows,
            schema=[
                "tab",
                "source_workbook",
                "workbook_sha256",
                "n_records",
                "n_reconstructed_labels",
                "extracted_at",
                "out_file",
            ],
            orient="row",
        ).sort("tab")
        manifest_path = args.out_dir / "manifest.csv"
        manifest_df.write_csv(manifest_path)
        print(f"wrote {manifest_path} ({manifest_df.height} records)")

        all_reconstructed: set[str] = set()
        for labels in tab_reconstructed.values():
            all_reconstructed.update(labels)
        print(
            f"reconstructed {len(all_reconstructed)} distance-bin label(s): "
            f"{sorted(all_reconstructed)}"
        )

        return 0
    finally:
        wb_values.close()
        wb_formulas.close()


if __name__ == "__main__":
    raise SystemExit(main())
