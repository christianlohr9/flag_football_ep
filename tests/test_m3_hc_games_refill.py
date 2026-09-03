"""Tests for `scripts/hc_games_refill.py` -- deterministic regeneration of
`data/reference/hc_games.csv`'s numeric-block-game rows.

Every fixture is built in-process with `openpyxl.Workbook()`; no test opens
the real, gitignored raw workbook directory. Synthetic workbooks are named after the two REAL declared
workbooks (`Offense Analytics 2026 Camps and Competitions.xlsx`,
`Scoring Probability by Situation 2023-2026.xlsx`) so `build_rows` exercises
the real `DECLARED_SOURCES` mapping, not a mocked one.
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import polars as pl
import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT / "scripts") not in sys.path:
    sys.path.insert(0, str(REPO_ROOT / "scripts"))

import hc_games_refill as refill  # noqa: E402

from flag_football_ep.reference import _HC_GAMES_SCHEMA  # noqa: E402

_OFFENSE_ANALYTICS_NAME = "Offense Analytics 2026 Camps and Competitions.xlsx"
_SCORING_PROBABILITY_NAME = "Scoring Probability by Situation 2023-2026.xlsx"


def _make_workbook(path: Path, sheets: dict[str, list[list]]) -> Path:
    wb = openpyxl.Workbook()
    default = wb.active
    first = True
    for name, rows in sheets.items():
        ws = default if first else wb.create_sheet(name)
        if first:
            ws.title = name
            first = False
        for row in rows:
            ws.append(row)
    wb.save(path)
    return path


_HEADER = ["PLAY #", "ODK", "DN", "DIST", "YARD LN", "RESULT", "GN/LS"]


def _empty_hc_games(tmp_path: Path) -> pl.DataFrame:
    return pl.DataFrame(schema=_HC_GAMES_SCHEMA)


def _write_hc_games_csv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    columns = list(_HC_GAMES_SCHEMA.keys())
    header = ",".join(columns)
    lines = [header]
    for row in rows:
        lines.append(",".join(str(row.get(c, "") or "") for c in columns))
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_competition_tier_csv(path: Path, rows: list[tuple[str, str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = ["source,competition,tier"] + [f"{s},{c},{t}" for s, c, t in rows]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


# --- build_rows: numeric-only emission, pair skipped -------------------------


def test_build_rows_two_block_sheet_emits_only_numeric_games(tmp_path: Path) -> None:
    numeric_game1 = [[i, "O", 1, 10, 25, "Rush", 5] for i in range(1, 6)]  # 5 plays
    numeric_game2 = [[i, "O", 1, 10, 25, "Complete", 5] for i in range(1, 8)]  # 7 plays
    pair_game = [["Germany", "Ireland", 1, 10, 25, "Rush", 5] for _ in range(6)]
    rows = [_HEADER, *numeric_game1, *numeric_game2, *pair_game]
    path = _make_workbook(tmp_path / _OFFENSE_ANALYTICS_NAME, {"Data": rows})

    existing = _empty_hc_games(tmp_path)
    new_rows, skips = refill.build_rows(path, "Data", existing)

    assert len(new_rows) == 2  # only the two numeric games
    assert all(r["game_id"].startswith("hc-") for r in new_rows)
    assert len(skips) == 1
    assert skips[0][3].startswith("pair block excluded by design")


def test_build_rows_game_id_unique_and_stable_across_two_runs(tmp_path: Path) -> None:
    numeric_game1 = [[i, "O", 1, 10, 25, "Rush", 5] for i in range(1, 6)]
    numeric_game2 = [[i, "O", 1, 10, 25, "Complete", 5] for i in range(1, 7)]
    rows = [_HEADER, *numeric_game1, *numeric_game2]
    path = _make_workbook(tmp_path / _OFFENSE_ANALYTICS_NAME, {"Data": rows})
    existing = _empty_hc_games(tmp_path)

    run1, _ = refill.build_rows(path, "Data", existing)
    run2, _ = refill.build_rows(path, "Data", existing)

    ids1 = [r["game_id"] for r in run1]
    ids2 = [r["game_id"] for r in run2]
    assert len(ids1) == len(set(ids1))  # unique within one run
    assert ids1 == ids2  # stable across runs


def test_build_rows_competition_is_locked_string(tmp_path: Path) -> None:
    numeric_game = [[i, "O", 1, 10, 25, "Rush", 5] for i in range(1, 6)]
    rows = [_HEADER, *numeric_game]
    path = _make_workbook(tmp_path / _OFFENSE_ANALYTICS_NAME, {"Data": rows})
    existing = _empty_hc_games(tmp_path)

    new_rows, _ = refill.build_rows(path, "Data", existing)

    assert new_rows[0]["competition"] == "HC Camps 2026"


def test_build_rows_scoring_probability_uses_its_own_locked_competition(tmp_path: Path) -> None:
    numeric_game = [[i, "O", 1, 10, 25, "Rush", 5] for i in range(1, 6)]
    rows = [_HEADER, *numeric_game]
    path = _make_workbook(tmp_path / _SCORING_PROBABILITY_NAME, {"Data": rows})
    existing = _empty_hc_games(tmp_path)

    new_rows, _ = refill.build_rows(path, "Data", existing)

    assert new_rows[0]["competition"] == "HC Charting 2023-2026"


# --- MIN_PLAYS ---------------------------------------------------------------


def test_build_rows_below_min_plays_skipped_and_left_provisional(tmp_path: Path) -> None:
    tiny_game = [[i, "O", 1, 10, 25, "Rush", 5] for i in range(1, 4)]  # 3 < MIN_PLAYS
    rows = [_HEADER, *tiny_game]
    path = _make_workbook(tmp_path / _OFFENSE_ANALYTICS_NAME, {"Data": rows})
    existing = _empty_hc_games(tmp_path)

    new_rows, skips = refill.build_rows(path, "Data", existing)

    assert new_rows == []
    assert len(skips) == 1
    assert "below MIN_PLAYS" in skips[0][3]


# --- preservation of existing declared rows -----------------------------------


def test_build_rows_existing_declared_row_preserved_not_regenerated(tmp_path: Path) -> None:
    numeric_game1 = [[i, "O", 1, 10, 25, "Rush", 5] for i in range(1, 6)]
    numeric_game2 = [[i, "O", 1, 10, 25, "Complete", 5] for i in range(1, 7)]
    rows = [_HEADER, *numeric_game1, *numeric_game2]
    path = _make_workbook(tmp_path / _OFFENSE_ANALYTICS_NAME, {"Data": rows})

    existing_games_path = tmp_path / "hc_games.csv"
    _write_hc_games_csv(
        existing_games_path,
        [
            {
                "workbook": "offense-analytics-2026-camps-and-competitions",
                "sheet": "data", "block_key": "b00-g00",
                "game_id": "hc-hand-declared-game", "home_team": "GER", "away_team": "OPP",
                "competition": "HC Camps 2026", "season": "2026", "tier": "mixed-other",
                "note": "hand-declared before refill",
            }
        ],
    )
    from flag_football_ep.reference import load_hc_games

    existing = load_hc_games(existing_games_path)

    new_rows, _ = refill.build_rows(path, "Data", existing)

    # only the SECOND game is new -- the first (b00-g00) is already declared
    assert len(new_rows) == 1
    assert new_rows[0]["block_key"] == "b00-g01"
    assert all(r["game_id"] != "hc-hand-declared-game" for r in new_rows)


def test_build_rows_missing_declared_key_raises_named(tmp_path: Path) -> None:
    """A declared (workbook, sheet, block_key) that no longer appears among
    the freshly segmented slices raises, naming it (RESEARCH Pitfall 2)."""
    numeric_game = [[i, "O", 1, 10, 25, "Rush", 5] for i in range(1, 6)]
    rows = [_HEADER, *numeric_game]  # only ONE game this time
    path = _make_workbook(tmp_path / _OFFENSE_ANALYTICS_NAME, {"Data": rows})

    existing_games_path = tmp_path / "hc_games.csv"
    _write_hc_games_csv(
        existing_games_path,
        [
            {
                "workbook": "offense-analytics-2026-camps-and-competitions",
                "sheet": "data", "block_key": "b00-g05",  # never produced by this fixture
                "game_id": "hc-stale-declared-game", "home_team": "GER", "away_team": "OPP",
                "competition": "HC Camps 2026", "season": "2026", "tier": "mixed-other",
                "note": "stale",
            }
        ],
    )
    from flag_football_ep.reference import load_hc_games

    existing = load_hc_games(existing_games_path)

    with pytest.raises(ValueError, match="b00-g05"):
        refill.build_rows(path, "Data", existing)


# --- undeclared (workbook, sheet) pair is a no-op -----------------------------


def test_build_rows_undeclared_workbook_returns_nothing(tmp_path: Path) -> None:
    numeric_game = [[i, "O", 1, 10, 25, "Rush", 5] for i in range(1, 6)]
    rows = [_HEADER, *numeric_game]
    path = _make_workbook(tmp_path / "Some Other Workbook.xlsx", {"Data": rows})
    existing = _empty_hc_games(tmp_path)

    new_rows, skips = refill.build_rows(path, "Data", existing)

    assert new_rows == []
    assert skips == []


# --- tier coverage assertion --------------------------------------------------


def test_assert_tier_coverage_raises_on_missing_row() -> None:
    rows = [
        {"workbook": "offense-analytics-2026-camps-and-competitions", "sheet": "data",
         "competition": "HC Camps 2026"},
    ]
    empty_tier = pl.DataFrame(schema={"source": pl.Utf8, "competition": pl.Utf8, "tier": pl.Utf8})

    with pytest.raises(ValueError, match="HC Camps 2026"):
        refill._assert_tier_coverage(rows, empty_tier)


def test_assert_tier_coverage_passes_when_row_present() -> None:
    rows = [
        {"workbook": "offense-analytics-2026-camps-and-competitions", "sheet": "data",
         "competition": "HC Camps 2026"},
    ]
    tier = pl.DataFrame(
        {
            "source": ["hc_workbook:offense-analytics-2026-camps-and-competitions:data"],
            "competition": ["HC Camps 2026"],
            "tier": ["mixed-other"],
        }
    )

    refill._assert_tier_coverage(rows, tier)  # must not raise


# --- main(): --dry-run writes nothing -----------------------------------------


def test_main_dry_run_prints_counts_and_writes_nothing(tmp_path: Path, capsys) -> None:
    numeric_game1 = [[i, "O", 1, 10, 25, "Rush", 5] for i in range(1, 6)]
    numeric_game2 = [[i, "O", 1, 10, 25, "Complete", 5] for i in range(1, 7)]
    pair_game = [["Germany", "Ireland", 1, 10, 25, "Rush", 5] for _ in range(6)]
    data_rows = [_HEADER, *numeric_game1, *numeric_game2, *pair_game]
    hc_dir = tmp_path / "data" / "input" / "hc_files"
    hc_dir.mkdir(parents=True)
    _make_workbook(hc_dir / _OFFENSE_ANALYTICS_NAME, {"Data": data_rows})

    ref_dir = tmp_path / "data" / "reference"
    hc_games_path = ref_dir / "hc_games.csv"
    _write_hc_games_csv(hc_games_path, [])
    tier_path = ref_dir / "competition_tier.csv"
    _write_competition_tier_csv(
        tier_path,
        [
            (
                "hc_workbook:offense-analytics-2026-camps-and-competitions:data",
                "HC Camps 2026",
                "mixed-other",
            ),
            (
                "hc_workbook:scoring-probability-by-situation-2023-2026:data",
                "HC Charting 2023-2026",
                "mixed-other",
            ),
        ],
    )

    toml_path = tmp_path / "ffep.toml"
    toml_path.write_text(
        f"""
[paths]
data_root = "data"
raw_hudl = "data/input/hudl"
raw_sportapp = "data/input/sportapp"
raw_ifaf = "data/input/ifaf"
raw_legacy = "data/input/legacy"
raw_hc_files = "data/input/hc_files"
processed = "data/processed"
reference = "data/reference"
models = "models"
mlruns = "mlruns"
contract = "docs/data-contract.schema.json"
reports = "reports"
video = "data/video"
labels = "data/labels"
tracking = "data/processed/tracking"

[reference]
half_boundaries = "data/reference/half_boundaries.csv"
final_scores = "data/reference/final_scores.csv"
team_mapping = "data/reference/team_mapping.csv"
sportapp_games = "data/reference/sportapp_games.csv"
competition_tier = "data/reference/competition_tier.csv"
player_mapping = "data/reference/player_mapping.csv"
group_opponents = "data/reference/group_opponents.csv"
hover_positions = "data/reference/hover_positions.csv"
homography_calibration = "data/reference/homography_calibration.csv"
gt_positions = "data/reference/gt_positions.csv"
continuity_review = "data/reference/continuity_review.csv"
hc_games = "data/reference/hc_games.csv"

[sources.sportapp]
base_url = "https://example.invalid"
api_key_env = "SPORTAPP_API_KEY"

[sources.ifaf]
base_url = "https://example.invalid"
tournament = "test"
api_key_env = "CPX_API_KEY"

[train]
exclude_games_ep = []
exclude_games_wp = []
ep_experiment = "ep_test"
wp_experiment = "wp_test"

[report]
own_team = "GER"
cycle_start_season = 2026

[cv]
pilot_session_id = "test"
detector_model = "test"
detector_experiment = "test"
resolution = 640
sahi = false
sahi_slice = 640
sahi_overlap = 0.2
train_epochs = 1
train_batch_size = 1
train_grad_accum = 1
device = "cpu"
label_frame_target = 1
cvat_host = "http://example.invalid"
cvat_username_env = "CVAT_USER"
cvat_password_env = "CVAT_PASS"
field_length_yards = 70.0
field_width_yards = 25.0
endzone_yards = 10.0
dvc_remote_name = "test"
dvc_remote_url = "test"
dvc_remote_endpoint = "test"
otc_obs_access_key_env = "OBS_KEY"
otc_obs_secret_key_env = "OBS_SECRET"
""",
        encoding="utf-8",
    )

    rc = refill.main(["--config", str(toml_path), "--dry-run"])

    assert rc == 0
    out = capsys.readouterr().out
    assert "2 new" in out
    assert "[dry-run] nothing written" in out
    assert not hc_games_path.read_text(encoding="utf-8").strip().endswith(",")  # unchanged (header-only)
    # the file was never overwritten -- still just the header line
    assert hc_games_path.read_text(encoding="utf-8").count("\n") <= 1
