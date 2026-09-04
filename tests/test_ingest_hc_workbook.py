"""Tests for HC workbook ingest: sheet reading, dtype-based block segmentation
(Task 2), and per-block contract mapping (Task 3).

Every fixture is built in-process with `openpyxl.Workbook()` using synthetic
team labels (`Alphaland`, `Betaland`) and synthetic player labels
(`Spieler A`, jersey `7`) -- no test may open a real gitignored HC workbook
or reference any real player/team name.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

import polars as pl

from flag_football_ep.canonical import CANONICAL_COLUMNS, NULLABLE_EXTRAS
from flag_football_ep.ingest.hc_workbook import (
    HcBlock,
    HcGameIdentity,
    HcGameSlice,
    SheetNotFoundError,
    count_result_tokens,
    hc_source_label,
    ingest_workbook,
    map_block_to_frame,
    read_sheet_rows,
    resolve_game_identity,
    segment_blocks,
    segment_games,
    slugify,
)
from flag_football_ep.reference import load_hc_games
from flag_football_ep.validation.checks import half_assigned
from flag_football_ep.validation.schema import HeaderReport, load_contract


def _make_workbook(tmp_path: Path, sheets: dict[str, list[list]]) -> Path:
    """Write a real .xlsx with one sheet per (name, rows) pair; rows[0] is the header."""
    wb = openpyxl.Workbook()
    default = wb.active
    first = True
    for name, rows in sheets.items():
        ws = default if first else wb.create_sheet(name)
        if first:
            ws.title = name
            first = False
        for row in rows:
            ws.append(row)
    path = tmp_path / "hc_test_workbook.xlsx"
    wb.save(path)
    return path


# --- slugify / hc_source_label ---------------------------------------------


def test_slugify_lowercases_and_collapses_non_alnum_runs() -> None:
    assert slugify("Offense Analytics 2026 Camps and Competitions") == (
        "offense-analytics-2026-camps-and-competitions"
    )
    assert slugify("Copy of Data") == "copy-of-data"


def test_hc_source_label_data_sheet() -> None:
    path = Path("Offense Analytics 2026 Camps and Competitions.xlsx")
    assert hc_source_label(path, "Data") == (
        "hc_workbook:offense-analytics-2026-camps-and-competitions:data"
    )


def test_hc_source_label_copy_of_data_sheet() -> None:
    path = Path("Offense Analytics 2026 Camps and Competitions.xlsx")
    assert hc_source_label(path, "Copy of Data") == (
        "hc_workbook:offense-analytics-2026-camps-and-competitions:copy-of-data"
    )


# --- read_sheet_rows ---------------------------------------------------------


def test_read_sheet_rows_skips_blank_rows_and_counts_them(tmp_path: Path) -> None:
    header = ["PLAY #", "ODK", "RESULT"]
    rows = [
        header,
        [1, "O", "Complete"],
        [None, None, None],
        [2, "O", "Incomplete"],
    ]
    path = _make_workbook(tmp_path, {"Data": rows})

    header_out, kept_rows, messages = read_sheet_rows(path, "Data")

    assert list(header_out) == header
    assert len(kept_rows) == 2
    assert kept_rows[0][0] == 2  # physical row number
    assert kept_rows[1][0] == 4
    joined = " ".join(messages)
    assert "1" in joined
    assert "übersprungen" in joined or "blank" in joined


def test_read_sheet_rows_empty_sheet_reports_leer_with_counts(tmp_path: Path) -> None:
    header = ["PLAY #", "ODK", "RESULT"]
    rows = [header] + [[None, None, None] for _ in range(200)]
    path = _make_workbook(tmp_path, {"Data": rows})

    _, kept_rows, messages = read_sheet_rows(path, "Data")

    assert kept_rows == []
    joined = " ".join(messages)
    assert "leer" in joined or "empty" in joined
    assert "200" in joined
    assert "0 Datenzeile" in joined


def test_read_sheet_rows_counts_na_cells(tmp_path: Path) -> None:
    header = ["PLAY #", "ODK", "RESULT"]
    rows = [
        header,
        [1, "O", "#N/A"],
        [2, "O", "Complete"],
    ]
    path = _make_workbook(tmp_path, {"Data": rows})

    _, _, messages = read_sheet_rows(path, "Data")

    assert any("#N/A" in m for m in messages)


def test_read_sheet_rows_missing_sheet_raises_with_available_names(tmp_path: Path) -> None:
    path = _make_workbook(tmp_path, {"Data": [["PLAY #", "ODK"], [1, "O"]]})

    with pytest.raises(SheetNotFoundError) as exc_info:
        read_sheet_rows(path, "Copy of Data")

    message = str(exc_info.value)
    assert "Copy of Data" in message
    assert "Data" in message


# --- segment_blocks -----------------------------------------------------------


def test_block_segmentation_pair_then_numeric(tmp_path: Path) -> None:
    header = ["PLAY #", "ODK", "OFF FORM"]
    pair_rows = [["Alphaland", "Betaland", "DOG"] for _ in range(5)]
    numeric_rows = [[i, "O", "DOG"] for i in range(1, 8)]
    rows = [header] + pair_rows + numeric_rows
    path = _make_workbook(tmp_path, {"Data": rows})

    header_out, kept_rows, _ = read_sheet_rows(path, "Data")
    blocks, messages = segment_blocks(header_out, kept_rows)

    assert len(blocks) == 2
    assert isinstance(blocks[0], HcBlock)
    assert blocks[0].kind == "pair"
    assert blocks[0].index == 0
    assert len(blocks[0].rows) == 5
    assert blocks[1].kind == "numeric"
    assert blocks[1].index == 1
    assert len(blocks[1].rows) == 7
    assert blocks[0].last_row < blocks[1].first_row
    assert any("Block 0" in m for m in messages)
    assert any("Block 1" in m for m in messages)


def test_block_segmentation_float_first_cell_is_numeric(tmp_path: Path) -> None:
    header = ["PLAY #", "ODK"]
    rows = [header, [1.0, "O"], [2.0, "O"]]
    path = _make_workbook(tmp_path, {"Data": rows})

    header_out, kept_rows, _ = read_sheet_rows(path, "Data")
    blocks, _ = segment_blocks(header_out, kept_rows)

    assert len(blocks) == 1
    assert blocks[0].kind == "numeric"


def test_block_segmentation_boolean_first_cell_not_numeric(tmp_path: Path) -> None:
    header = ["PLAY #", "ODK"]
    rows = [header, [True, "O"]]
    path = _make_workbook(tmp_path, {"Data": rows})

    header_out, kept_rows, _ = read_sheet_rows(path, "Data")
    blocks, _ = segment_blocks(header_out, kept_rows)

    assert len(blocks) == 0


def test_block_segmentation_skips_empty_first_cell_without_new_block(tmp_path: Path) -> None:
    header = ["PLAY #", "ODK", "RESULT"]
    rows = [
        header,
        [1, "O", "Complete"],
        ["", None, "orphan note"],
        [2, "O", "Incomplete"],
    ]
    path = _make_workbook(tmp_path, {"Data": rows})

    header_out, kept_rows, _ = read_sheet_rows(path, "Data")
    blocks, _ = segment_blocks(header_out, kept_rows)

    assert len(blocks) == 1
    assert blocks[0].kind == "numeric"
    assert len(blocks[0].rows) == 2


def test_block_segmentation_twenty_consecutive_numeric_rows_is_one_block(tmp_path: Path) -> None:
    header = ["PLAY #", "ODK"]
    rows = [header] + [[i, "O"] for i in range(1, 21)]
    path = _make_workbook(tmp_path, {"Data": rows})

    header_out, kept_rows, _ = read_sheet_rows(path, "Data")
    blocks, _ = segment_blocks(header_out, kept_rows)

    assert len(blocks) == 1
    assert len(blocks[0].rows) == 20


# --- map_block_to_frame (Task 3) --------------------------------------------

# Offense Analytics 2026 Camps and Competitions.xlsx :: Data (M3-01-RESEARCH.md
# "Verified workbook facts") -- one clean numeric block, no pair rows at all.
OFFENSE_ANALYTICS_HEADER = [
    "PLAY #", "ODK", "OFF FORM", "OFF STR", "OFF PLAY", "DN", "DIST", "YARD LN",
    "RESULT", "GN/LS", "TARGET ROUTE", "RECEIVED BY", "AIR YARDS", "Hand",
    "Efficiency", "Thrown By", "Target", "X", "S", "C", "Q", "Y", "Drop", "B",
]

# Scoring Probability by Situation 2023-2026.xlsx :: Data -- the workbook
# with team-name-pair rows; note GN/LS sits AFTER RECEIVED BY here (unlike
# Offense Analytics), matching M3-01-RESEARCH.md Pitfall 2's finding that
# GN/LS is itself part of the pair block's unresolved tail in this workbook.
SCORING_PROBABILITY_HEADER = [
    "PLAY #", "ODK", "OFF FORM", "OFF STR", "OFF PLAY", "DN", "DIST", "YARD LN",
    "RESULT", "Drive Success", "TARGET ROUTE", "RECEIVED BY", "GN/LS", "Thrown By",
    "YAC", "QB", "C", "X/H", "Y/CAT", "Z", "Target", "Drop", "B",
]


@pytest.fixture
def contract(contract_path: Path):
    return load_contract(contract_path)


def _numeric_block(header: list, data_rows: list[list]) -> HcBlock:
    rows = [(i + 2, tuple(row)) for i, row in enumerate(data_rows)]
    return HcBlock(
        index=0, kind="numeric", header=header, rows=rows,
        first_row=rows[0][0], last_row=rows[-1][0],
    )


def _pair_block(header: list, data_rows: list[list]) -> HcBlock:
    rows = [(i + 2, tuple(row)) for i, row in enumerate(data_rows)]
    return HcBlock(
        index=0, kind="pair", header=header, rows=rows,
        first_row=rows[0][0], last_row=rows[-1][0],
    )


_OFFENSE_ROW = [
    1.0, "O", "DOG", "RIGHT", "SWEEP", 1.0, 10.0, -20.0, "Rush", 5.0,
    "SLANT", "Spieler A", 3.0, "R", 1.0, "Spieler B", "7", "", "", "", "", "", "", "",
]


def test_header_and_block_mapping_numeric_block_materializes_play_type(contract) -> None:
    block = _numeric_block(OFFENSE_ANALYTICS_HEADER, [_OFFENSE_ROW])

    df, _, _, messages = map_block_to_frame(block, contract)

    for name in ("PLAY #", "ODK", "DN", "DIST", "YARD LN", "RESULT", "GN/LS"):
        assert name in df.columns
    assert "PLAY TYPE" in df.columns
    assert df["PLAY TYPE"].null_count() == df.height
    assert any("PLAY TYPE" in m for m in messages)


def test_header_and_block_mapping_validate_header_never_raises_numeric(contract) -> None:
    block = _numeric_block(OFFENSE_ANALYTICS_HEADER, [_OFFENSE_ROW])

    df, header_report, _, _ = map_block_to_frame(block, contract)  # must not raise

    assert isinstance(header_report, HeaderReport)
    assert header_report.missing_core == []


def test_header_and_block_mapping_validate_header_never_raises_pair(contract) -> None:
    pair_row = [
        "Alphaland", "Betaland", "DOG", "RIGHT", "SWEEP", 1.0, 10.0, -20.0,
        "Rush", 1.0, "SLANT", "Spieler A", 5.0, "Spieler B", 3.0, "7",
        "", "", "", "", "", "", "",
    ]
    block = _pair_block(SCORING_PROBABILITY_HEADER, [pair_row])

    df, header_report, _, _ = map_block_to_frame(block, contract)  # must not raise

    assert header_report.missing_core == []


def test_pair_block_tail_nulled_with_notice(contract) -> None:
    pair_row = [
        "Alphaland", "Betaland", "DOG", "RIGHT", "SWEEP", 1.0, 10.0, -20.0,
        "Rush", 1.0, "SLANT", "Spieler A", 5.0, "Spieler B", 3.0, "7",
        "", "", "", "", "", "", "",
    ]
    block = _pair_block(SCORING_PROBABILITY_HEADER, [pair_row])

    df, _, _, messages = map_block_to_frame(block, contract)

    # columns through TARGET ROUTE carry real values
    assert df["target_route"][0] == "SLANT"
    assert df["OFF FORM"][0] == "DOG"
    # RECEIVED BY onward is null (renamed to received_by); GN/LS sits after
    # RECEIVED BY in this workbook's real header, so it is null here too
    assert df["received_by"].null_count() == 1
    assert df["GN/LS"].null_count() == 1
    # PLAY #/ODK are null for pair-block rows, but the raw team-pair values
    # survive under dedicated names for plan M3-01-03's game-identity key
    assert df["PLAY #"].null_count() == 1
    assert df["ODK"].null_count() == 1
    assert df["hc_pair_team1"][0] == "Alphaland"
    assert df["hc_pair_team2"][0] == "Betaland"
    joined = " ".join(messages)
    assert "RECEIVED BY" in joined
    assert "Frage 2" in joined or "offen" in joined


def test_pair_block_odk_derived_from_marker_row_tail_still_nulled(contract) -> None:
    """A marker row's ODK is the row's own O/D/S value (Frage 2, Antwort
    2026-09-03) -- but the tail from RECEIVED BY onward stays nulled for a
    marker row exactly like a header row: Frage 2's answer addressed the
    block-segmentation/ODK question, not the tail-column question."""
    header_row = [
        "Alphaland", "Betaland", "DOG", "RIGHT", "SWEEP", 1.0, 10.0, -20.0,
        "Rush", 1.0, "SLANT", "Spieler A", 5.0, "Spieler B", 3.0, "7",
        "", "", "", "", "", "", "",
    ]
    marker_row = [
        "O", None, "DOG", "RIGHT", "SWEEP", 2.0, 5.0, -15.0,
        "Complete", 1.0, "SLANT", "Spieler C", 8.0, "Spieler D", 3.0, "7",
        "", "", "", "", "", "", "",
    ]
    block = _pair_block(SCORING_PROBABILITY_HEADER, [header_row, marker_row])

    df, _, _, messages = map_block_to_frame(block, contract)

    assert df["PLAY #"].null_count() == 2
    assert df["ODK"].to_list() == [None, "O"]
    assert df["received_by"].null_count() == 2
    joined = " ".join(messages)
    assert "1 Marker-Zeile(n)" in joined
    assert "1 Zeile(n)" in joined  # the header row's own count phrase


def test_jersey_string_integral_float_becomes_plain_integer_string(contract) -> None:
    block = _numeric_block(OFFENSE_ANALYTICS_HEADER, [_OFFENSE_ROW])

    df, _, _, _ = map_block_to_frame(block, contract)

    assert df["air_yards"][0] == "3"
    assert df["PLAY #"][0] == "1"


def test_drop_column_in_canonical_columns_as_nullable_utf8_extra() -> None:
    assert "drop" in CANONICAL_COLUMNS
    assert NULLABLE_EXTRAS["drop"] == pl.Utf8


@pytest.mark.parametrize("drop_header", ["Drop", "DROP", "drop", " Drop "])
def test_drop_header_any_case_or_whitespace_maps_to_drop_column_raw_text(
    contract, drop_header
) -> None:
    header = list(OFFENSE_ANALYTICS_HEADER)
    header[22] = drop_header  # "Drop" sits at index 22 in the real header
    row = list(_OFFENSE_ROW)
    row[22] = "X"  # the head coach's real charted drop mark
    block = _numeric_block(header, [row])

    df, _, _, messages = map_block_to_frame(block, contract)

    assert df["drop"][0] == "X"
    # "drop" is a rename target, not a header collected into the unmapped
    # ("nicht stillschweigend verworfen") notice.
    unmapped_notices = [m for m in messages if "nicht stillschweigend verworfen" in m]
    assert not unmapped_notices or drop_header not in unmapped_notices[0]


def test_result_negative_float_preserved_dn_out_of_range_flagged(contract) -> None:
    row = list(_OFFENSE_ROW)
    row[5] = 7.0  # DN out of the contract's [0, 4] range
    row[8] = -5.0  # RESULT: the real data-entry error from the corpus
    block = _numeric_block(OFFENSE_ANALYTICS_HEADER, [row])

    df, _, domain_violations, _ = map_block_to_frame(block, contract)

    assert df["RESULT"][0] == "-5.0"
    dn_violations = [v for v in domain_violations if v.column == "DN" and v.rule == "range"]
    assert dn_violations, f"expected a DN range violation, got {domain_violations}"


def test_unknown_headers_named_not_silently_dropped(contract) -> None:
    block = _numeric_block(OFFENSE_ANALYTICS_HEADER, [_OFFENSE_ROW])

    _, _, _, messages = map_block_to_frame(block, contract)

    joined = " ".join(messages)
    for unmapped_header in ("X", "S", "C", "Q", "Y", "B"):
        assert unmapped_header in joined
    # Drop is mapped to the canonical `drop` extra (M3-04-06) -- no longer
    # among the unmapped-header notices.
    assert "Drop" not in joined


def test_unknown_headers_charting_extras_are_renamed_not_listed(contract) -> None:
    block = _numeric_block(OFFENSE_ANALYTICS_HEADER, [_OFFENSE_ROW])

    df, _, _, _ = map_block_to_frame(block, contract)

    for extra in ("air_yards", "hand", "efficiency", "target_route", "received_by", "thrown_by", "target", "drop"):
        assert extra in df.columns


def test_header_dedup_appends_suffix_and_names_it(contract) -> None:
    header = list(OFFENSE_ANALYTICS_HEADER)
    header[18] = "C"  # duplicate of the existing "C" at index 19 -> "C" / "C_2"
    block = _numeric_block(header, [_OFFENSE_ROW])

    df, _, _, messages = map_block_to_frame(block, contract)

    assert "C" in df.columns
    assert "C_2" in df.columns
    assert any("C_2" in m for m in messages)


# --- segment_games (Task 2) --------------------------------------------------


def _hc_games_frame(rows: list[dict]) -> pl.DataFrame:
    """Build an in-memory `hc_games` frame matching `_HC_GAMES_SCHEMA`'s columns/dtypes."""
    columns = [
        "workbook", "sheet", "block_key", "source_team1", "source_team2", "game_id",
        "home_team", "away_team", "competition", "season", "game_date", "tier",
        "corpus_game_id", "note",
    ]

    def _season_value(row: dict) -> int | None:
        raw = row.get("season")
        return int(raw) if raw not in (None, "") else None

    data = {
        col: ([_season_value(row) for row in rows] if col == "season" else [row.get(col) for row in rows])
        for col in columns
    }
    schema = {col: (pl.Int32 if col == "season" else pl.Utf8) for col in columns}
    return pl.DataFrame(data, schema=schema)


def test_game_segmentation_numeric_twenty_consecutive_rows_is_one_game() -> None:
    header = ["PLAY #", "ODK"]
    rows = [(i + 2, (float(i), "O")) for i in range(1, 21)]
    block = HcBlock(index=0, kind="numeric", header=header, rows=rows, first_row=2, last_row=21)

    slices, _ = segment_games(block)

    assert len(slices) == 1
    assert len(slices[0].rows) == 20
    assert slices[0].block_key == "b00-g00"


def test_game_segmentation_numeric_play_number_reset_splits_two_games() -> None:
    header = ["PLAY #", "ODK"]
    first_game = [(i + 2, (float(i), "O")) for i in range(1, 9)]  # PLAY # 1..8
    second_game = [(i + 10, (float(i), "O")) for i in range(1, 13)]  # PLAY # 1..12, reset
    rows = first_game + second_game
    block = HcBlock(
        index=0, kind="numeric", header=header, rows=rows,
        first_row=rows[0][0], last_row=rows[-1][0],
    )

    slices, _ = segment_games(block)

    assert len(slices) == 2
    assert len(slices[0].rows) == 8
    assert len(slices[1].rows) == 12
    assert slices[0].block_key == "b00-g00"
    assert slices[1].block_key == "b00-g01"


def test_game_segmentation_numeric_null_play_number_forces_boundary() -> None:
    header = ["PLAY #", "ODK"]
    rows = [
        (2, (1.0, "O")),
        (3, (2.0, "O")),
        (4, (None, "O")),  # unparseable -- forces a boundary
        (5, (4.0, "O")),
    ]
    block = HcBlock(index=0, kind="numeric", header=header, rows=rows, first_row=2, last_row=5)

    slices, messages = segment_games(block)

    assert len(slices) >= 2  # the null row starts a new game
    joined = " ".join(messages)
    assert "PLAY #" in joined


def test_game_segmentation_pair_block_splits_on_team_pair_change() -> None:
    header = ["PLAY #", "ODK", "RESULT"]
    rows = [
        (2, ("Alphaland", "Betaland", "Rush")),
        (3, ("Alphaland", "Betaland", "Complete")),
        (4, ("Gammaland", "Deltaland", "Rush")),
    ]
    block = HcBlock(index=1, kind="pair", header=header, rows=rows, first_row=2, last_row=4)

    slices, _ = segment_games(block)

    assert len(slices) == 2
    assert len(slices[0].rows) == 2
    assert slices[0].source_team1 == "Alphaland"
    assert slices[0].source_team2 == "Betaland"
    assert len(slices[1].rows) == 1
    assert slices[1].source_team1 == "Gammaland"
    assert slices[1].source_team2 == "Deltaland"
    assert slices[0].block_key == "b01-g00"
    assert slices[1].block_key == "b01-g01"


def test_game_segmentation_pair_block_case_and_whitespace_insensitive_match() -> None:
    header = ["PLAY #", "ODK"]
    rows = [
        (2, ("Alphaland", "Betaland")),
        (3, (" alphaland ", " BETALAND ")),
        (4, ("Gammaland", "Deltaland")),
    ]
    block = HcBlock(index=0, kind="pair", header=header, rows=rows, first_row=2, last_row=4)

    slices, _ = segment_games(block)

    assert len(slices) == 2
    assert len(slices[0].rows) == 2
    # raw label from the first row of the slice is preserved verbatim
    assert slices[0].source_team1 == "Alphaland"


def test_game_segmentation_block_key_zero_padded_and_block_scoped() -> None:
    header = ["PLAY #", "ODK"]
    rows = [(i + 2, (float(i), "O")) for i in range(1, 3)] + [
        (i + 20, (float(i), "O")) for i in range(1, 3)
    ]
    block = HcBlock(
        index=3, kind="numeric", header=header, rows=rows,
        first_row=rows[0][0], last_row=rows[-1][0],
    )

    slices, _ = segment_games(block)

    assert [s.block_key for s in slices] == ["b03-g00", "b03-g01"]


def test_game_segmentation_empty_block_returns_no_slices() -> None:
    block = HcBlock(index=0, kind="numeric", header=["PLAY #"], rows=[], first_row=0, last_row=0)

    slices, messages = segment_games(block)

    assert slices == []
    assert messages == []


def test_game_segmentation_pair_block_possession_swap_is_one_game() -> None:
    """Mirrors the real Data-tab pattern (M3-02-RESEARCH.md Sec 1.2): the head
    coach charts offense and defense possessions of the SAME game as flipped
    team-pair rows. An unordered-pair boundary key must group all five rows
    into one slice, not fragment on every possession flip."""
    header = ["PLAY #", "ODK", "RESULT"]
    rows = [
        (2, ("Germany", "Ireland", "Rush")),
        (3, ("Germany", "Ireland", "Complete")),
        (4, ("Ireland", "Germany", "Rush")),
        (5, ("Ireland", "Germany", "Complete")),
        (6, ("Germany", "Ireland", "Rush")),
    ]
    block = HcBlock(index=1, kind="pair", header=header, rows=rows, first_row=2, last_row=6)

    slices, _ = segment_games(block)

    assert len(slices) == 1
    assert len(slices[0].rows) == 5
    # RAW labels of the first row survive verbatim, not normalized/sorted
    assert slices[0].source_team1 == "Germany"
    assert slices[0].source_team2 == "Ireland"


def test_game_segmentation_pair_block_possession_swap_then_real_opponent_change_splits() -> None:
    """A genuine opponent change after a possession-swap stretch still opens
    a new slice."""
    header = ["PLAY #", "ODK", "RESULT"]
    rows = [
        (2, ("Germany", "Ireland", "Rush")),
        (3, ("Germany", "Ireland", "Complete")),
        (4, ("Ireland", "Germany", "Rush")),
        (5, ("Ireland", "Germany", "Complete")),
        (6, ("Germany", "Ireland", "Rush")),
        (7, ("Germany", "Spain", "Rush")),
    ]
    block = HcBlock(index=1, kind="pair", header=header, rows=rows, first_row=2, last_row=7)

    slices, _ = segment_games(block)

    assert len(slices) == 2
    assert len(slices[0].rows) == 5
    assert len(slices[1].rows) == 1
    assert slices[1].source_team1 == "Germany"
    assert slices[1].source_team2 == "Spain"


def test_game_segmentation_pair_block_possession_swap_case_and_whitespace_insensitive() -> None:
    """Case/whitespace insensitivity survives the unordered-pair comparison:
    a swapped, differently-cased/whitespace-padded pair does not open a new
    slice."""
    header = ["PLAY #", "ODK"]
    rows = [
        (2, ("Germany", "Ireland")),
        (3, (" ireland ", "GERMANY")),
        (4, ("Germany", "Ireland")),
    ]
    block = HcBlock(index=0, kind="pair", header=header, rows=rows, first_row=2, last_row=4)

    slices, _ = segment_games(block)

    assert len(slices) == 1
    assert len(slices[0].rows) == 3


def test_game_segmentation_pair_block_single_row_noise_stays_its_own_slice() -> None:
    """A single-row noise entry with an unmatched abbreviation between two
    possession-swap stretches becomes its own one-row slice -- never merged
    into either neighbour by inference (RESEARCH Sec 1.2: do not guess
    ambiguous abbreviations like S/F)."""
    header = ["PLAY #", "ODK"]
    rows = [
        (2, ("Germany", "Ireland")),
        (3, ("Ireland", "Germany")),
        (4, ("AT", "D")),
        (5, ("Germany", "Ireland")),
        (6, ("Ireland", "Germany")),
    ]
    block = HcBlock(index=0, kind="pair", header=header, rows=rows, first_row=2, last_row=6)

    slices, _ = segment_games(block)

    assert len(slices) == 3
    assert len(slices[0].rows) == 2
    assert len(slices[1].rows) == 1
    assert slices[1].source_team1 == "AT"
    assert slices[1].source_team2 == "D"
    assert len(slices[2].rows) == 2


def test_game_segmentation_pair_block_three_slices_block_key_scoped() -> None:
    """block_key numbering stays b{block:02d}-g{game:02d} and block-scoped
    across a three-slice pair block."""
    header = ["PLAY #", "ODK"]
    rows = [
        (2, ("Germany", "Ireland")),
        (3, ("Ireland", "Germany")),
        (4, ("AT", "D")),
        (5, ("Germany", "Ireland")),
    ]
    block = HcBlock(index=1, kind="pair", header=header, rows=rows, first_row=2, last_row=5)

    slices, _ = segment_games(block)

    assert [s.block_key for s in slices] == ["b01-g00", "b01-g01", "b01-g02"]


def test_game_segmentation_numeric_block_unaffected_by_pair_block_change() -> None:
    """Numeric-block segmentation (PLAY#-reset) is untouched by the pair-block
    unordered-key change."""
    header = ["PLAY #", "ODK"]
    first_game = [(i + 2, (float(i), "O")) for i in range(1, 9)]
    second_game = [(i + 10, (float(i), "O")) for i in range(1, 13)]
    rows = first_game + second_game
    block = HcBlock(
        index=0, kind="numeric", header=header, rows=rows,
        first_row=rows[0][0], last_row=rows[-1][0],
    )

    slices, _ = segment_games(block)

    assert len(slices) == 2
    assert len(slices[0].rows) == 8
    assert len(slices[1].rows) == 12


# --- pair-block header + O/D/S-marker rule (M3-02-04 deviation, Frage 2 ------
# Antwort 2026-09-03: "O" = Offense, "D" = Defense, "S" = no-play; a
# team-name row opens a block, O/D/S rows inherit it, a blank row or the
# next team-name row closes it) --------------------------------------------


def test_game_segmentation_pair_block_marker_rows_inherit_open_header() -> None:
    header = ["PLAY #", "ODK", "RESULT"]
    rows = [
        (2, ("Germany", "Ireland", "Rush")),  # header row
        (3, ("O", None, "Complete")),  # marker row: offense
        (4, ("D", None, "Rush")),  # marker row: defense
    ]
    block = HcBlock(index=0, kind="pair", header=header, rows=rows, first_row=2, last_row=4)

    slices, messages = segment_games(block)

    assert len(slices) == 1
    assert len(slices[0].rows) == 3
    assert slices[0].source_team1 == "Germany"
    assert slices[0].source_team2 == "Ireland"
    assert "ohne Team-Namenspaar-Kopfzeile" not in " ".join(messages)


def test_game_segmentation_pair_block_marker_case_and_whitespace_insensitive() -> None:
    header = ["PLAY #", "ODK"]
    rows = [
        (2, ("Germany", "Ireland")),
        (3, (" o ", None)),
        (4, ("d", None)),
        (5, ("S", None)),
    ]
    block = HcBlock(index=0, kind="pair", header=header, rows=rows, first_row=2, last_row=5)

    slices, _ = segment_games(block)

    assert len(slices) == 1
    assert len(slices[0].rows) == 4


def test_game_segmentation_pair_block_row_number_gap_alone_is_not_a_boundary() -> None:
    """A gap in physical row numbers -- which happens both for a genuinely
    blank row (stripped by read_sheet_rows) AND for any other row
    segment_blocks silently skips (e.g. a populated play row with an empty
    column A -- 5 such rows found in the real `Data`-tab pair block,
    2026-09-03) -- is NOT, by itself, treated as a boundary: the two causes
    are indistinguishable at this layer, and an earlier version of this
    function that inferred a boundary from the gap over-fragmented the real
    block (137 -> 18 instead of the validated 137 -> 22, M3-02-RESEARCH.md
    Sec 1.2). A marker row after a gap still inherits the last-seen header,
    exactly as if there had been no gap."""
    header = ["PLAY #", "ODK"]
    rows = [
        (2, ("Germany", "Ireland")),
        (3, ("O", None)),
        # row 4 skipped for some reason not visible to this function
        (5, ("D", None)),
    ]
    block = HcBlock(index=0, kind="pair", header=header, rows=rows, first_row=2, last_row=5)

    slices, _ = segment_games(block)

    assert len(slices) == 1
    assert len(slices[0].rows) == 3
    assert slices[0].source_team1 == "Germany"
    assert slices[0].source_team2 == "Ireland"


def test_game_segmentation_pair_block_new_header_closes_marker_block_even_same_pair() -> None:
    """Once a block has received a marker row, a repeated team-name row
    (even the SAME pair, no gap) always opens a new block -- the head
    coach's literal rule ('bis ... einer neuen Zeile mit Teamnamen'), not
    the possession-swap merge rule (which only applies to a pure
    team-name-per-row stretch that never saw a marker)."""
    header = ["PLAY #", "ODK"]
    rows = [
        (2, ("Germany", "Ireland")),
        (3, ("O", None)),
        (4, ("Germany", "Ireland")),  # new header row, no gap
        (5, ("D", None)),
    ]
    block = HcBlock(index=0, kind="pair", header=header, rows=rows, first_row=2, last_row=5)

    slices, _ = segment_games(block)

    assert len(slices) == 2
    assert len(slices[0].rows) == 2
    assert len(slices[1].rows) == 2


def test_game_segmentation_pair_block_marker_rows_before_any_header_are_headerless() -> None:
    header = ["PLAY #", "ODK"]
    rows = [
        (2, ("O", None)),
        (3, ("D", None)),
        (4, ("Germany", "Ireland")),
        (5, ("O", None)),
    ]
    block = HcBlock(index=0, kind="pair", header=header, rows=rows, first_row=2, last_row=5)

    slices, messages = segment_games(block)

    assert len(slices) == 2
    assert slices[0].source_team1 is None
    assert slices[0].source_team2 is None
    assert len(slices[0].rows) == 2
    assert slices[1].source_team1 == "Germany"
    assert len(slices[1].rows) == 2
    joined = " ".join(messages)
    assert "1 Pair-Block-Gruppe(n) ohne Team-Namenspaar-Kopfzeile" in joined
    assert "docs/hc-blocks-ohne-kopfzeile.md" in joined


def test_game_segmentation_pair_block_era1_and_marker_style_coexist_without_cross_talk() -> None:
    """A block mixing a pure team-name-per-row (possession-swap) stretch and
    a header+marker stretch, separated by a genuine opponent change, keeps
    each style's own rule -- no cross-contamination between the two."""
    header = ["PLAY #", "ODK"]
    rows = [
        (2, ("Germany", "Ireland")),  # era-1 possession swap
        (3, ("Ireland", "Germany")),
        (4, ("Germany", "Ireland")),
        (5, ("Germany", "Spain")),  # header, marker era
        (6, ("O", None)),
        (7, ("D", None)),
    ]
    block = HcBlock(index=0, kind="pair", header=header, rows=rows, first_row=2, last_row=7)

    slices, _ = segment_games(block)

    assert len(slices) == 2
    assert len(slices[0].rows) == 3
    assert slices[0].source_team1 == "Germany"
    assert slices[0].source_team2 == "Ireland"
    assert len(slices[1].rows) == 3
    assert slices[1].source_team1 == "Germany"
    assert slices[1].source_team2 == "Spain"


# --- resolve_game_identity (Task 2) ------------------------------------------


def test_resolve_game_identity_mapped_hit_returns_hc_games_row() -> None:
    hc_games = _hc_games_frame(
        [
            {
                "workbook": "offense-analytics-2026", "sheet": "data", "block_key": "b00-g00",
                "source_team1": "", "source_team2": "", "game_id": "hc-2026-01-alp-bet",
                "home_team": "ALP", "away_team": "BET", "competition": "Camp",
                "season": "2026", "game_date": "2026-01-10", "tier": "womens-national",
                "corpus_game_id": "", "note": "",
            }
        ]
    )
    header = ["PLAY #", "ODK"]
    rows = [(2, (1.0, "O")), (3, (2.0, "O"))]
    slice_ = HcGameSlice(
        block_index=0, game_index=0, block_key="b00-g00", kind="numeric",
        rows=rows, first_row=2, last_row=3, source_team1=None, source_team2=None,
    )

    identity, messages = resolve_game_identity(
        slice_, "offense-analytics-2026", "data", hc_games
    )

    assert isinstance(identity, HcGameIdentity)
    assert identity.provisional is False
    assert identity.game_id == "hc-2026-01-alp-bet"
    assert identity.home_team == "ALP"
    assert identity.away_team == "BET"
    assert identity.tier == "womens-national"
    assert identity.season == 2026
    assert messages == []


def test_resolve_game_identity_same_block_key_different_sheet_resolves_differently() -> None:
    hc_games = _hc_games_frame(
        [
            {
                "workbook": "wb", "sheet": "data", "block_key": "b00-g00",
                "source_team1": "", "source_team2": "", "game_id": "hc-data-game",
                "home_team": "ALP", "away_team": "BET", "competition": "Camp",
                "season": "2026", "game_date": "2026-01-10", "tier": "womens-national",
                "corpus_game_id": "", "note": "",
            },
            {
                "workbook": "wb", "sheet": "copy-of-data", "block_key": "b00-g00",
                "source_team1": "", "source_team2": "", "game_id": "hc-copy-game",
                "home_team": "GAM", "away_team": "DEL", "competition": "Camp",
                "season": "2026", "game_date": "2026-01-11", "tier": "womens-national",
                "corpus_game_id": "", "note": "",
            },
        ]
    )
    slice_ = HcGameSlice(
        block_index=0, game_index=0, block_key="b00-g00", kind="numeric",
        rows=[(2, (1.0, "O"))], first_row=2, last_row=2, source_team1=None, source_team2=None,
    )

    identity_data, _ = resolve_game_identity(slice_, "wb", "data", hc_games)
    identity_copy, _ = resolve_game_identity(slice_, "wb", "copy-of-data", hc_games)

    assert identity_data.game_id == "hc-data-game"
    assert identity_copy.game_id == "hc-copy-game"


def test_provisional_game_miss_returns_provisional_id_and_notice() -> None:
    hc_games = _hc_games_frame([])
    slice_ = HcGameSlice(
        block_index=0, game_index=0, block_key="b00-g00", kind="numeric",
        rows=[(2, (1.0, "O")), (3, (2.0, "O"))], first_row=2, last_row=3,
        source_team1=None, source_team2=None,
    )

    identity, messages = resolve_game_identity(slice_, "offense-analytics-2026", "data", hc_games)

    assert identity.provisional is True
    assert identity.game_id == "hc-offense-analytics-2026-data-b00-g00"
    assert len(messages) == 1
    message = messages[0]
    assert "b00-g00" in message
    assert "2" in message  # row range / play count present
    assert "3" in message


def test_provisional_game_pair_block_notice_includes_raw_team_labels() -> None:
    hc_games = _hc_games_frame([])
    slice_ = HcGameSlice(
        block_index=1, game_index=0, block_key="b01-g00", kind="pair",
        rows=[(2, ("Alphaland", "Betaland", "Rush"))], first_row=2, last_row=2,
        source_team1="Alphaland", source_team2="Betaland",
    )

    identity, messages = resolve_game_identity(slice_, "wb", "data", hc_games)

    assert identity.provisional is True
    message = messages[0]
    assert "Alphaland" in message
    assert "Betaland" in message


def test_provisional_game_carries_null_fields_nothing_invented() -> None:
    hc_games = _hc_games_frame([])
    slice_ = HcGameSlice(
        block_index=0, game_index=0, block_key="b00-g00", kind="numeric",
        rows=[(2, (1.0, "O"))], first_row=2, last_row=2, source_team1=None, source_team2=None,
    )

    identity, _ = resolve_game_identity(slice_, "wb", "data", hc_games)

    assert identity.home_team is None
    assert identity.away_team is None
    assert identity.tier is None
    assert identity.season is None
    assert identity.game_date is None


def test_provisional_game_hc_games_csv_round_trip(tmp_path: Path) -> None:
    """One CSV round-trip through load_hc_games proves segment_games/resolve_game_identity
    fit the real loader, not just the inline test frame builder."""
    path = tmp_path / "hc_games.csv"
    path.write_text(
        "workbook,sheet,block_key,source_team1,source_team2,game_id,home_team,away_team,"
        "competition,season,game_date,tier,corpus_game_id,note\n"
        "offense-analytics-2026,data,b00-g00,,,hc-2026-01-alp-bet,ALP,BET,Camp,2026,"
        "2026-01-10,womens-national,,\n",
        encoding="utf-8",
    )
    hc_games = load_hc_games(path)

    slice_ = HcGameSlice(
        block_index=0, game_index=0, block_key="b00-g00", kind="numeric",
        rows=[(2, (1.0, "O"))], first_row=2, last_row=2, source_team1=None, source_team2=None,
    )

    identity, messages = resolve_game_identity(slice_, "offense-analytics-2026", "data", hc_games)

    assert identity.provisional is False
    assert identity.game_id == "hc-2026-01-alp-bet"
    assert messages == []


# --- count_result_tokens (Task 3) --------------------------------------------


def test_count_result_tokens_counts_every_token() -> None:
    df = pl.DataFrame({"RESULT": ["Rush", "Complete, TD", "Rush", None, "Block"]})

    counts = count_result_tokens(df)

    assert counts == {"Rush": 2, "Complete": 1, "TD": 1, "Block": 1}


def test_count_result_tokens_empty_frame_returns_empty_dict() -> None:
    df = pl.DataFrame({"RESULT": []}, schema={"RESULT": pl.Utf8})

    assert count_result_tokens(df) == {}


# --- ingest_workbook (Task 3) -------------------------------------------------

_INGEST_MINIMAL_HEADER = ["PLAY #", "ODK", "DN", "DIST", "YARD LN", "RESULT", "GN/LS"]

_INGEST_FULL_HEADER = [
    "PLAY #", "ODK", "OFF FORM", "DN", "DIST", "YARD LN", "RESULT", "GN/LS",
    "RECEIVED BY", "QB", "THROWN BY", "TARGET", "TACKLE",
]


def _player_mapping_frame(rows: list[tuple[str, str, str]]) -> pl.DataFrame:
    return pl.DataFrame(
        rows, schema=["source", "source_player", "canonical_player"], orient="row"
    )


def test_ingest_workbook_two_game_numeric_block_resolves_mapped_game_ids(
    tmp_path: Path, contract
) -> None:
    rows = [
        _INGEST_MINIMAL_HEADER,
        [1, "O", 1, 10, -20, "Rush", 5],
        [2, "O", 2, 5, -15, "Complete", 5],
        [1, "O", 1, 10, -20, "Rush", 5],  # PLAY # resets -> second game
    ]
    path = _make_workbook(tmp_path, {"Data": rows})

    hc_games = _hc_games_frame(
        [
            {
                "workbook": "hc-test-workbook", "sheet": "data", "block_key": "b00-g00",
                "source_team1": "", "source_team2": "", "game_id": "hc-g1",
                "home_team": "ALP", "away_team": "BET", "competition": "Camp",
                "season": "2026", "game_date": "2026-01-01", "tier": "womens-national",
                "corpus_game_id": "", "note": "",
            },
            {
                "workbook": "hc-test-workbook", "sheet": "data", "block_key": "b00-g01",
                "source_team1": "", "source_team2": "", "game_id": "hc-g2",
                "home_team": "GAM", "away_team": "DEL", "competition": "Camp",
                "season": "2026", "game_date": "2026-01-02", "tier": "womens-national",
                "corpus_game_id": "", "note": "",
            },
        ]
    )
    player_mapping = _player_mapping_frame([])

    df, notices = ingest_workbook(path, "Data", contract, hc_games, player_mapping)

    assert list(df.columns) == list(CANONICAL_COLUMNS)
    assert df["source"].unique().to_list() == [hc_source_label(path, "Data")]
    assert sorted(df["game_id"].unique().to_list()) == ["hc-g1", "hc-g2"]
    assert notices.sheet == "Data"


def test_ingest_workbook_no_drop_header_materializes_null_and_reports_it(
    tmp_path: Path, contract
) -> None:
    """A sheet with no Drop header still conforms -- `drop` is a typed null,
    named in `ConformReport.materialized_extras`, same as the four M3-01-02
    extras behave when their header is absent."""
    rows = [
        _INGEST_MINIMAL_HEADER,
        [1, "O", 1, 10, -20, "Rush", 5],
    ]
    path = _make_workbook(tmp_path, {"Data": rows})
    hc_games = _hc_games_frame([])
    player_mapping = _player_mapping_frame([])

    df, notices = ingest_workbook(path, "Data", contract, hc_games, player_mapping)

    assert "drop" in df.columns
    assert df["drop"].null_count() == df.height
    assert "drop" in notices.conform.materialized_extras


def test_ingest_workbook_drop_header_present_survives_to_canonical_frame(
    tmp_path: Path, contract
) -> None:
    header = _INGEST_MINIMAL_HEADER + ["Drop"]
    rows = [
        header,
        [1, "O", 1, 10, -20, "Rush", 5, "X"],
        [2, "O", 2, 5, -15, "Incomplete", 5, ""],
    ]
    path = _make_workbook(tmp_path, {"Data": rows})
    hc_games = _hc_games_frame([])
    player_mapping = _player_mapping_frame([])

    df, notices = ingest_workbook(path, "Data", contract, hc_games, player_mapping)

    # an empty charted cell reads back as null (openpyxl), not empty string
    assert df.sort("play_id")["drop"].to_list() == ["X", None]
    assert "drop" not in notices.conform.materialized_extras


def test_ingest_workbook_empty_sheet_returns_zero_row_canonical_frame(
    tmp_path: Path, contract
) -> None:
    header = _INGEST_MINIMAL_HEADER
    rows = [header] + [[None] * len(header) for _ in range(50)]
    path = _make_workbook(tmp_path, {"Data": rows})

    hc_games = _hc_games_frame([])
    player_mapping = _player_mapping_frame([])

    df, notices = ingest_workbook(path, "Data", contract, hc_games, player_mapping)  # must not raise

    assert df.height == 0
    assert list(df.columns) == list(CANONICAL_COLUMNS)
    assert len(notices.messages) >= 1


def test_ingest_workbook_end_to_end_two_sheet_workbook(tmp_path: Path, contract) -> None:
    numeric_game1 = [
        [1, "O", "DOG", 1, 10, -20, "Rush", 5, None, "Spieler A", None, None, None],
        [2, "O", "DOG", 2, 5, -15, "Complete", 5, "25", "Spieler A", "Spieler A", "25", None],
        [3, "D", "DOG", 3, 3, -10, "Sack", 0, None, "Spieler A", None, None, "Spieler A"],
    ]
    numeric_game2 = [
        [1, "O", "DOG", 1, 10, -20, "Complete, TD", 20, "Spieler A", "Spieler A", "Spieler A", "Spieler A", None],
    ]
    pair_game = [
        ["Alphaland", "Betaland", "DOG", 1, 10, -20, "Rush", 5, "Spieler B", None, None, None, None],
        ["Alphaland", "Betaland", "DOG", 1, 8, -18, "Complete", 8, "Spieler C", None, None, None, None],
    ]
    data_rows = [_INGEST_FULL_HEADER] + numeric_game1 + numeric_game2 + pair_game
    copy_rows = [_INGEST_FULL_HEADER] + numeric_game1

    path = _make_workbook(tmp_path, {"Data": data_rows, "Copy of Data": copy_rows})

    hc_games = _hc_games_frame(
        [
            {
                "workbook": "hc-test-workbook", "sheet": "data", "block_key": "b00-g00",
                "source_team1": "", "source_team2": "", "game_id": "hc-2026-01-alp-bet",
                "home_team": "ALP", "away_team": "BET", "competition": "Test Camp",
                "season": "2026", "game_date": "2026-01-05", "tier": "womens-national",
                "corpus_game_id": "", "note": "",
            }
        ]
    )
    player_mapping = _player_mapping_frame([("hc_workbook", "Spieler A", "Anna Mustermann")])

    df_data, notices_data = ingest_workbook(path, "Data", contract, hc_games, player_mapping)
    df_copy, notices_copy = ingest_workbook(
        path, "Copy of Data", contract, hc_games, player_mapping
    )

    # canonical column set
    assert list(df_data.columns) == list(CANONICAL_COLUMNS)
    assert list(df_copy.columns) == list(CANONICAL_COLUMNS)

    # per-sheet source label, uniform for every row
    assert df_data["source"].unique().to_list() == [hc_source_label(path, "Data")]
    assert df_copy["source"].unique().to_list() == [hc_source_label(path, "Copy of Data")]
    assert notices_data.source_label != notices_copy.source_label

    # per-game ids: mapped game1, provisional game2 and provisional pair game
    game_ids = set(df_data["game_id"].unique().to_list())
    assert "hc-2026-01-alp-bet" in game_ids
    assert any(g.startswith("hc-hc-test-workbook-data-b00-g01") for g in game_ids)
    assert any(g.startswith("hc-hc-test-workbook-data-b01-g00") for g in game_ids)

    # null pair-block posteam/defteam
    pair_rows = df_data.filter(pl.col("game_id").str.starts_with("hc-hc-test-workbook-data-b01"))
    assert pair_rows.height == 2
    assert pair_rows["posteam"].null_count() == 2
    assert pair_rows["defteam"].null_count() == 2

    # token counts include every RESULT token seen, including the multi-token combo
    assert notices_data.result_token_counts.get("Rush", 0) >= 2
    assert notices_data.result_token_counts.get("Complete", 0) >= 2
    assert notices_data.result_token_counts.get("TD", 0) == 1
    assert notices_data.result_token_counts.get("Sack", 0) == 1

    # unmapped-player behaviour: "Spieler A" mapped, jersey "25" left unmapped.
    # "Spieler B"/"Spieler C" (pair-block RECEIVED BY values) never reach
    # map_players at all: RECEIVED BY onward is nulled for pair-block rows
    # (PAIR_BLOCK_TAIL_ANCHOR, Frage 2 undetermined column shift) -- their
    # raw labels are dropped before player mapping even runs, the strongest
    # possible PII-safety property for that column.
    assert "Spieler A" not in notices_data.unmapped_players
    assert "25" in notices_data.unmapped_players
    assert "Spieler B" not in notices_data.unmapped_players
    assert "Spieler C" not in notices_data.unmapped_players
    joined = " ".join(notices_data.messages)
    # PII discipline: only the count appears in the human-readable message, never a label
    assert "Spieler B" not in joined
    assert "Spieler C" not in joined
    assert str(len(notices_data.unmapped_players)) in joined


def test_ingest_workbook_player_identity_mixed_type_survives_mapping(
    tmp_path: Path, contract
) -> None:
    rows = [
        _INGEST_FULL_HEADER,
        [1, "O", "DOG", 1, 10, -20, "Complete", 5, "25", "Spieler A", "Spieler A", "25", None],
        [2, "O", "DOG", 2, 5, -15, "Complete", 5, "Spieler A", "Spieler A", "Spieler A", "Spieler A", None],
    ]
    path = _make_workbook(tmp_path, {"Data": rows})

    hc_games = _hc_games_frame([])
    player_mapping = _player_mapping_frame([("hc_workbook", "Spieler A", "Anna Mustermann")])

    df, notices = ingest_workbook(path, "Data", contract, hc_games, player_mapping)  # must not raise

    received_by = df["received_by"].to_list()
    assert "25" in received_by  # unmapped jersey label survives verbatim
    assert "Anna Mustermann" in received_by  # mapped name label replaced
    assert "25" in notices.unmapped_players
    assert "Anna Mustermann" not in notices.unmapped_players
    assert "Spieler A" not in notices.unmapped_players


def test_ingest_workbook_synthesized_play_id_for_pair_block_rows(tmp_path: Path, contract) -> None:
    rows = [
        _INGEST_MINIMAL_HEADER,
        ["Alphaland", "Betaland", 1, 10, -20, "Rush", 5],
        ["Alphaland", "Betaland", 1, 8, -18, "Complete", 8],
        ["Alphaland", "Betaland", 1, 6, -16, "Complete", 6],
    ]
    path = _make_workbook(tmp_path, {"Data": rows})

    hc_games = _hc_games_frame([])
    player_mapping = _player_mapping_frame([])

    df, notices = ingest_workbook(path, "Data", contract, hc_games, player_mapping)

    assert sorted(df["play_id"].to_list()) == [1, 2, 3]
    joined = " ".join(notices.messages)
    assert "3" in joined
    assert "neu vergeben" in joined
    assert "hc_play_no" in joined


# --- HC corpus admission rule 1 (placeholder rows, 2026-09-04) ---


def test_ingest_workbook_placeholder_rows_removed_before_validation(
    tmp_path: Path, contract
) -> None:
    """A numeric block's placeholder rows (ODK/DN/RESULT all null -- the two
    PLAY # 1-2 rows charted at the start of a new SP-charted game, before
    anything real happens) are removed before validation ever runs (HC
    corpus admission rule 1, confirmed 2026-09-04)."""
    rows = [
        _INGEST_MINIMAL_HEADER,
        [1, None, None, None, None, None, None],
        [2, None, None, None, None, None, None],
        [3, "O", 1, 10, -20, "Rush", 5],
    ]
    path = _make_workbook(tmp_path, {"Data": rows})

    hc_games = _hc_games_frame(
        [
            {
                "workbook": "hc-test-workbook", "sheet": "data", "block_key": "b00-g00",
                "source_team1": "", "source_team2": "", "game_id": "hc-g1",
                "home_team": "ALP", "away_team": "BET", "competition": "Camp",
                "season": "2026", "game_date": "2026-01-01", "tier": "womens-national",
                "corpus_game_id": "", "note": "",
            },
        ]
    )
    player_mapping = _player_mapping_frame([])

    df, notices = ingest_workbook(path, "Data", contract, hc_games, player_mapping)

    assert df.height == 1
    assert df["down"].to_list() == [1]
    joined = " ".join(notices.messages)
    assert "2 Platzhalter-Zeile(n)" in joined
    assert "hc-g1" in joined


def test_ingest_workbook_placeholder_rows_keep_real_result_null_dn(
    tmp_path: Path, contract
) -> None:
    """A row with a real RESULT but null DN is a genuinely, if incompletely,
    charted play -- it stays and is checked as before (not a placeholder)."""
    rows = [
        _INGEST_MINIMAL_HEADER,
        [1, "O", None, None, None, "Rush", 5],
    ]
    path = _make_workbook(tmp_path, {"Data": rows})

    hc_games = _hc_games_frame(
        [
            {
                "workbook": "hc-test-workbook", "sheet": "data", "block_key": "b00-g00",
                "source_team1": "", "source_team2": "", "game_id": "hc-g1",
                "home_team": "ALP", "away_team": "BET", "competition": "Camp",
                "season": "2026", "game_date": "2026-01-01", "tier": "womens-national",
                "corpus_game_id": "", "note": "",
            },
        ]
    )
    player_mapping = _player_mapping_frame([])

    df, notices = ingest_workbook(path, "Data", contract, hc_games, player_mapping)

    assert df.height == 1
    joined = " ".join(notices.messages)
    assert "Platzhalter-Zeile(n)" not in joined


# --- half sentinel ---


def test_ingest_workbook_declared_game_gets_half_sentinel(tmp_path: Path, contract) -> None:
    """A declared, non-`Copy of Data` game's rows all carry half=2 (Int32)
    and PASS half_assigned (M3-02-RESEARCH Sec 2.2: 2 is the only sentinel
    that satisfies half_assigned AND keeps game_end/Winner/No_Score correct)."""
    rows = [
        _INGEST_MINIMAL_HEADER,
        [1, "O", 1, 10, -20, "Rush", 5],
        [2, "O", 2, 5, -15, "Complete", 5],
    ]
    path = _make_workbook(tmp_path, {"Data": rows})

    hc_games = _hc_games_frame(
        [
            {
                "workbook": "hc-test-workbook", "sheet": "data", "block_key": "b00-g00",
                "source_team1": "", "source_team2": "", "game_id": "hc-g1",
                "home_team": "ALP", "away_team": "BET", "competition": "Camp",
                "season": "2026", "game_date": "2026-01-01", "tier": "womens-national",
                "corpus_game_id": "", "note": "",
            },
        ]
    )
    player_mapping = _player_mapping_frame([])

    df, notices = ingest_workbook(path, "Data", contract, hc_games, player_mapping)

    assert df["half"].dtype == pl.Int32
    assert df["half"].to_list() == [2, 2]

    results = half_assigned(df)
    assert all(r.status.name == "PASS" for r in results)


def test_ingest_workbook_undeclared_game_stays_half_null_alongside_declared(
    tmp_path: Path, contract
) -> None:
    """A declared game (half=2) and an undeclared game (half=null) coexist
    row-wise in the same frame -- the decision is per game, not per sheet."""
    rows = [
        _INGEST_MINIMAL_HEADER,
        [1, "O", 1, 10, -20, "Rush", 5],  # declared game
        [1, "O", 1, 10, -20, "Rush", 5],  # PLAY # resets -> undeclared second game
    ]
    path = _make_workbook(tmp_path, {"Data": rows})

    hc_games = _hc_games_frame(
        [
            {
                "workbook": "hc-test-workbook", "sheet": "data", "block_key": "b00-g00",
                "source_team1": "", "source_team2": "", "game_id": "hc-g1",
                "home_team": "ALP", "away_team": "BET", "competition": "Camp",
                "season": "2026", "game_date": "2026-01-01", "tier": "womens-national",
                "corpus_game_id": "", "note": "",
            },
        ]
    )
    player_mapping = _player_mapping_frame([])

    df, notices = ingest_workbook(path, "Data", contract, hc_games, player_mapping)

    declared_half = df.filter(pl.col("game_id") == "hc-g1")["half"].to_list()
    undeclared_half = df.filter(pl.col("game_id") != "hc-g1")["half"].to_list()
    assert declared_half == [2]
    assert undeclared_half == [None]

    results = half_assigned(df)
    by_game = {r.game_id: r for r in results}
    assert by_game["hc-g1"].status.name == "PASS"
    undeclared_game_id = next(g for g in by_game if g != "hc-g1")
    assert by_game[undeclared_game_id].status.name == "FAIL"
    assert "no half boundary" in by_game[undeclared_game_id].detail


def test_ingest_workbook_copy_of_data_stays_half_null_even_if_declared(
    tmp_path: Path, contract
) -> None:
    """`Copy of Data` rows stay half=null even for a game declared in
    hc_games.csv -- the sheet exclusion overrides the declaration (Frage 2
    unresolved column layout, M3-02-RESEARCH Sec 1.3)."""
    rows = [
        _INGEST_MINIMAL_HEADER,
        [1, "O", 1, 10, -20, "Rush", 5],
        [2, "O", 2, 5, -15, "Complete", 5],
    ]
    path = _make_workbook(tmp_path, {"Copy of Data": rows})

    hc_games = _hc_games_frame(
        [
            {
                "workbook": "hc-test-workbook", "sheet": "copy-of-data", "block_key": "b00-g00",
                "source_team1": "", "source_team2": "", "game_id": "hc-g1",
                "home_team": "ALP", "away_team": "BET", "competition": "Camp",
                "season": "2026", "game_date": "2026-01-01", "tier": "womens-national",
                "corpus_game_id": "", "note": "",
            },
        ]
    )
    player_mapping = _player_mapping_frame([])

    df, notices = ingest_workbook(path, "Copy of Data", contract, hc_games, player_mapping)

    assert df["half"].dtype == pl.Int32
    assert df["half"].null_count() == df.height

    results = half_assigned(df)
    assert all(r.status.name == "FAIL" for r in results)
    assert all("no half boundary" in r.detail for r in results)


def test_ingest_workbook_half_sentinel_notices_carry_counts_and_reasons_no_pii(
    tmp_path: Path, contract
) -> None:
    """The two German notices name real counts and reasons
    (`nicht in hc_games.csv deklariert` / `Copy of Data, Frage 2 offen`);
    neither message contains a player name."""
    rows = [
        _INGEST_FULL_HEADER,
        [1, "O", "DOG", 1, 10, -20, "Rush", 5, None, "Spieler A", None, None, None],
        [1, "O", "DOG", 1, 10, -20, "Rush", 5, None, "Spieler A", None, None, None],
    ]
    path = _make_workbook(tmp_path, {"Data": rows, "Copy of Data": rows})

    hc_games = _hc_games_frame(
        [
            {
                "workbook": "hc-test-workbook", "sheet": "data", "block_key": "b00-g00",
                "source_team1": "", "source_team2": "", "game_id": "hc-g1",
                "home_team": "ALP", "away_team": "BET", "competition": "Camp",
                "season": "2026", "game_date": "2026-01-01", "tier": "womens-national",
                "corpus_game_id": "", "note": "",
            },
        ]
    )
    player_mapping = _player_mapping_frame([("hc_workbook", "Spieler A", "Anna Mustermann")])

    df_data, notices_data = ingest_workbook(path, "Data", contract, hc_games, player_mapping)
    df_copy, notices_copy = ingest_workbook(
        path, "Copy of Data", contract, hc_games, player_mapping
    )

    joined_data = " ".join(notices_data.messages)
    assert "nicht in hc_games.csv deklariert" in joined_data
    assert "Anna Mustermann" not in joined_data
    assert "Spieler A" not in joined_data

    joined_copy = " ".join(notices_copy.messages)
    assert "Copy of Data" in joined_copy
    assert "Frage 2 offen" in joined_copy
    assert "Anna Mustermann" not in joined_copy
    assert "Spieler A" not in joined_copy


def test_ingest_workbook_empty_sheet_half_column_stays_int32(tmp_path: Path, contract) -> None:
    """An empty sheet still returns the zero-row canonical frame with a
    `half` column of dtype Int32 -- schema is independent of any declared
    game."""
    header = _INGEST_MINIMAL_HEADER
    rows = [header] + [[None] * len(header) for _ in range(10)]
    path = _make_workbook(tmp_path, {"Data": rows})

    hc_games = _hc_games_frame([])
    player_mapping = _player_mapping_frame([])

    df, notices = ingest_workbook(path, "Data", contract, hc_games, player_mapping)

    assert df.height == 0
    assert df["half"].dtype == pl.Int32


# --- pair-block header + O/D/S-marker rule, end-to-end (M3-02-04 deviation) --


def test_ingest_workbook_declared_marker_rows_get_real_posteam_defteam(
    tmp_path: Path, contract
) -> None:
    """A declared pair-block game whose rows are a team-name header followed
    by O/D marker rows resolves real posteam/defteam for the marker rows --
    the header row itself still has no ODK, so it stays null (Frage 2,
    Antwort 2026-09-03 addressed the marker rows, not the header row)."""
    rows = [
        _INGEST_MINIMAL_HEADER,
        ["Alphaland", "Betaland", 1, 10, -20, "Rush", 5],
        ["O", None, 2, 5, -15, "Complete", 5],
        ["D", None, 1, 10, -20, "Rush", 5],
    ]
    path = _make_workbook(tmp_path, {"Data": rows})

    hc_games = _hc_games_frame(
        [
            {
                "workbook": "hc-test-workbook", "sheet": "data", "block_key": "b00-g00",
                "source_team1": "", "source_team2": "", "game_id": "hc-g1",
                "home_team": "ALP", "away_team": "BET", "competition": "Camp",
                "season": "2026", "game_date": "2026-01-01", "tier": "womens-national",
                "corpus_game_id": "", "note": "",
            },
        ]
    )
    player_mapping = _player_mapping_frame([])

    df, notices = ingest_workbook(path, "Data", contract, hc_games, player_mapping)
    df = df.sort("play_id")

    # PLAY # is null for both the header and the marker rows (synthesized
    # from row position, PLAY #/ODK's raw content: row 1 = header, row 2 =
    # "O" marker, row 3 = "D" marker) -- play_id 1/2/3 in sheet order.
    assert df["game_id"].unique().to_list() == ["hc-g1"]
    assert df["play_id"].to_list() == [1, 2, 3]
    header_row, offense_row, defense_row = df[0], df[1], df[2]
    assert header_row["posteam"].null_count() == 1
    assert header_row["defteam"].null_count() == 1
    assert offense_row["posteam"].to_list() == ["ALP"]
    assert offense_row["defteam"].to_list() == ["BET"]
    assert defense_row["posteam"].to_list() == ["BET"]
    assert defense_row["defteam"].to_list() == ["ALP"]


def test_ingest_workbook_pair_block_marker_s_becomes_no_play(tmp_path: Path, contract) -> None:
    """An `S` marker row (Frage 2 Antwort: 'S für no-play') gets
    play_type='no_play' regardless of its RESULT token -- consistent with
    Timeout/Offsetting Penalties handling (contract v1.2)."""
    rows = [
        _INGEST_MINIMAL_HEADER,
        ["Alphaland", "Betaland", 1, 10, -20, "Rush", 5],
        ["S", None, 0, None, None, "Rush", None],
    ]
    path = _make_workbook(tmp_path, {"Data": rows})

    hc_games = _hc_games_frame(
        [
            {
                "workbook": "hc-test-workbook", "sheet": "data", "block_key": "b00-g00",
                "source_team1": "", "source_team2": "", "game_id": "hc-g1",
                "home_team": "ALP", "away_team": "BET", "competition": "Camp",
                "season": "2026", "game_date": "2026-01-01", "tier": "womens-national",
                "corpus_game_id": "", "note": "",
            },
        ]
    )
    player_mapping = _player_mapping_frame([])

    df, notices = ingest_workbook(path, "Data", contract, hc_games, player_mapping)
    df = df.sort("play_id")

    s_row = df[1]  # play_id 2, the "S" marker row
    assert s_row["play_type"].to_list() == ["no_play"]
    joined = " ".join(notices.messages)
    assert "play_type='no_play' markiert" in joined


def test_ingest_workbook_headerless_marker_group_stays_provisional_no_pii(
    tmp_path: Path, contract
) -> None:
    """A marker-only run with no preceding team-name header row (Frage 2
    Antwort: 'wenn er irgendwann aufgehört hat, Teamnamen aufzuschreiben')
    resolves to a provisional identity -- `home_team`/`away_team` stay null
    (no declared team names known), but `posteam`/`defteam` still resolve
    from the real `ODK` marker via the HC-OFF/HC-DEF placeholders (HC
    corpus admission rule 3, confirmed 2026-09-04: EP only needs the
    offense/defense perspective, not real team identity). The
    headerless-group notice never contains a player label."""
    rows = [
        _INGEST_MINIMAL_HEADER,
        ["O", None, 1, 10, -20, "Rush", 5],
        ["D", None, 2, 5, -15, "Complete", 5],
    ]
    path = _make_workbook(tmp_path, {"Data": rows})

    hc_games = _hc_games_frame([])
    player_mapping = _player_mapping_frame([])

    df, notices = ingest_workbook(path, "Data", contract, hc_games, player_mapping)

    assert df.height == 2
    assert df["home_team"].null_count() == 2
    assert df["away_team"].null_count() == 2
    assert df["posteam"].to_list() == ["HC-OFF", "HC-DEF"]
    assert df["defteam"].to_list() == ["HC-DEF", "HC-OFF"]
    joined = " ".join(notices.messages)
    assert "1 Pair-Block-Gruppe(n) ohne Team-Namenspaar-Kopfzeile" in joined
    assert "docs/hc-blocks-ohne-kopfzeile.md" in joined
    assert "Spieler" not in joined
